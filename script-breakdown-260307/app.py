"""Flask web application for Script Breakdown (剧本拆解)."""
import csv
import io
import json
import math
import os
import tempfile
import threading
import urllib.parse
import zipfile
from datetime import datetime

from flask import Flask, render_template, request, jsonify, Response

from src.parsers.base import ParseResult
from src.parsers.txt_parser import TxtParser
from src.parsers.docx_parser import DocxParser
from src.parsers.pdf_parser import PdfParser
from src.parsers.fdx_parser import FdxParser
from src.scene.models import Scene
from src.scene.detector import SceneDetector
from src.scene.llm_detector import LLMSceneDetector
from src.scene.entity_extractor import extract_entities as rule_extract_entities
from src.scene.patterns import parse_heading_fields
from src.llm.base import LLMConfig
from src.llm.claude_adapter import ClaudeAdapter
from src.llm.openai_adapter import OpenAIAdapter
from src.llm.ollama_adapter import OllamaAdapter
from src.web.state import AppState
from src.web.project import Project, PROJECTS_DIR
from src.schedule.models import (
    ActorSchedule, LocationInfo, ScheduleConfig,
    ProductionSchedule, ScheduleSnapshot,
)
from src.schedule.optimizer import ScheduleOptimizer
from src.schedule.duration_engine import SceneDurationEngine, DurationParams
from src.schedule.constraints import ConstraintChecker
from src.schedule.evaluator import ScheduleEvaluator
from src.schedule.llm_loop import LLMScheduleLoop
from src.schedule.learner import SchedulePreferenceLearner

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB

# ── LLM Config Persistence ────────────────────────────────────────

_LLM_CONFIG_FILE = os.path.expanduser("~/.script-breakdown/llm_config.json")
_LAST_PROJECT_FILE = os.path.expanduser("~/.script-breakdown/last_project.json")

# Guard so auto-restore only runs once per server process
_project_init_done = False


def _save_llm_config(config: LLMConfig) -> None:
    os.makedirs(os.path.dirname(_LLM_CONFIG_FILE), exist_ok=True)
    with open(_LLM_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "provider": config.provider,
            "model_name": config.model_name,
            "api_key": config.api_key,
            "base_url": config.base_url,
            "temperature": config.temperature,
            "max_tokens": config.max_tokens,
        }, f, ensure_ascii=False)


def _load_llm_config() -> LLMConfig:
    if os.path.exists(_LLM_CONFIG_FILE):
        try:
            with open(_LLM_CONFIG_FILE, "r", encoding="utf-8") as f:
                d = json.load(f)
            return LLMConfig(
                provider=d.get("provider", ""),
                model_name=d.get("model_name", ""),
                api_key=d.get("api_key", ""),
                base_url=d.get("base_url", ""),
                temperature=d.get("temperature", 0.3),
                max_tokens=d.get("max_tokens", 4096),
            )
        except Exception:
            pass
    return LLMConfig()


def _save_last_project(project_id: str) -> None:
    """Persist the last opened project ID so it can be restored after restart."""
    os.makedirs(os.path.dirname(_LAST_PROJECT_FILE), exist_ok=True)
    with open(_LAST_PROJECT_FILE, "w", encoding="utf-8") as f:
        json.dump({"project_id": project_id}, f)


def _clear_last_project() -> None:
    if os.path.exists(_LAST_PROJECT_FILE):
        try:
            os.remove(_LAST_PROJECT_FILE)
        except OSError:
            pass


def _try_restore_last_project(state: AppState) -> None:
    """Auto-load the last opened project after a server restart."""
    if not os.path.exists(_LAST_PROJECT_FILE):
        return
    try:
        with open(_LAST_PROJECT_FILE, "r", encoding="utf-8") as f:
            d = json.load(f)
        project_id = d.get("project_id", "")
        if not project_id:
            return
        project = Project.load(project_id)
        state.project = project
        if project.meta.active_episode_id:
            try:
                state.load_from_episode(project.meta.active_episode_id)
            except Exception:
                pass
        state.status_message = f"已恢复项目: {project.meta.name}"
    except Exception:
        pass  # Silently ignore if project no longer exists


@app.before_request
def _init_state():
    """Load persisted config and auto-restore the last project on first request."""
    global _project_init_done
    state = AppState()
    if not state.llm_config.provider:
        state.llm_config = _load_llm_config()
    if not _project_init_done:
        _project_init_done = True
        if not state.project:
            _try_restore_last_project(state)

PARSERS = {
    ".txt": TxtParser,
    ".text": TxtParser,
    ".pdf": PdfParser,
    ".docx": DocxParser,
    ".fdx": FdxParser,
}

def _create_llm(config: LLMConfig):
    """Always use the universal OpenAIAdapter regardless of saved provider value.

    OpenAIAdapter auto-detects the API format from the base_url:
      • base_url contains 'anthropic.com' → Anthropic Messages API
      • everything else                   → OpenAI Chat Completions format
    Legacy 'claude' / 'ollama' provider values are silently upgraded.
    """
    return OpenAIAdapter(config)


# ── Page ─────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/overview")
def overview():
    return render_template("overview.html")


# ── Upload ───────────────────────────────────────────────────────

@app.route("/api/upload", methods=["POST"])
def upload():
    state = AppState()

    if "file" not in request.files:
        return jsonify({"error": "没有上传文件"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "文件名为空"}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    parser_cls = PARSERS.get(ext)
    if parser_cls is None:
        return jsonify({"error": f"不支持的文件格式: {ext}"}), 400

    # If 'add_as_episode=true' and a project is open, delegate to episode-add logic
    add_as_episode = request.form.get("add_as_episode", "false").lower() == "true"
    episode_name = request.form.get("episode_name", "").strip() or os.path.splitext(file.filename)[0]

    # Save to temp file for parser
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        file.save(tmp)
        tmp_path = tmp.name

    try:
        parser = parser_cls()
        parse_result = parser.parse(tmp_path)

        if add_as_episode and state.project:
            # Save current episode first
            state.save_to_episode()

            # Create new episode entry
            episode = state.project.add_episode(episode_name, file.filename)

            # Load new content into state
            state.parse_result = parse_result
            state.filename = file.filename
            state.scene_list = None
            state.entities.clear()
            state.global_entities = {"characters": [], "props": []}
            state.character_analyses.clear()
            state.global_analysis = ""
            state._undo_stack.clear()
            state._redo_stack.clear()
            state.current_episode_id = episode.id
            state.project.meta.active_episode_id = episode.id

            detector = SceneDetector()
            state.scene_list = detector.detect(state.parse_result)
            state.push_undo()
            state.save_to_episode()

            state.status_message = (
                f"已添加集数「{episode_name}」: {file.filename} "
                f"({len(parse_result.lines)} 行), 识别 {len(state.scene_list)} 个场次"
            )
        else:
            # Standalone import (no project context)
            state.reset()
            state.parse_result = parse_result
            state.filename = file.filename

            detector = SceneDetector()
            state.scene_list = detector.detect(state.parse_result)
            state.push_undo()

            state.status_message = (
                f"已导入: {file.filename} ({len(state.parse_result.lines)} 行), "
                f"自动识别 {len(state.scene_list)} 个场次"
            )

        return jsonify(state.to_dict())

    except Exception as e:
        return jsonify({"error": f"解析文件失败: {e}"}), 500
    finally:
        os.unlink(tmp_path)


# ── State ────────────────────────────────────────────────────────

@app.route("/api/state")
def get_state():
    state = AppState()
    return jsonify(state.to_dict())


@app.route("/api/scenes/global")
def get_scenes_global():
    """Return all scenes + entities across every episode with global sequential
    scene_numbers.  The schedule page uses this so _sceneMap / _entityMap use
    the same numbering as the schedule's scene_ids (produced by _collect_all_scenes).
    """
    state = AppState()
    all_scenes, entities_by_num = _collect_all_scenes(state)
    scenes_data = [
        {
            "scene_number": s.scene_number,
            "heading":      s.heading,
            "location":     s.location,
            "time_of_day":  s.time_of_day,
            "int_ext":      s.int_ext,
            "start_line":   s.start_line,
            "end_line":     s.end_line,
        }
        for s in all_scenes
    ]
    return jsonify({
        "scenes":   scenes_data,
        "entities": {str(k): v for k, v in entities_by_num.items()},
    })


# ── Undo / Redo ──────────────────────────────────────────────────

@app.route("/api/undo", methods=["POST"])
def undo():
    state = AppState()
    if state.undo():
        return jsonify(state.to_dict())
    return jsonify({"error": "没有可撤销的操作"}), 400


@app.route("/api/redo", methods=["POST"])
def redo():
    state = AppState()
    if state.redo():
        return jsonify(state.to_dict())
    return jsonify({"error": "没有可恢复的操作"}), 400


# ── Scene Update ─────────────────────────────────────────────────

@app.route("/api/scenes/<int:idx>", methods=["PUT"])
def update_scene(idx):
    state = AppState()
    if not state.scene_list or idx < 0 or idx >= len(state.scene_list):
        return jsonify({"error": "无效的场次索引"}), 400

    state.push_undo()

    scene = state.scene_list[idx]
    data = request.get_json()

    if "heading" in data:
        scene.heading = data["heading"]
    if "int_ext" in data:
        scene.int_ext = data["int_ext"]
    if "location" in data:
        scene.location = data["location"]
    if "time_of_day" in data:
        scene.time_of_day = data["time_of_day"]
    if "summary" in data:
        scene.summary = data["summary"]
    if "scene_type" in data:
        scene.scene_type = data["scene_type"]

    scene.is_manually_adjusted = True
    state.status_message = f"场次 {idx + 1} 已更新"

    return jsonify(state.to_dict())


# ── Reparse scene heading ─────────────────────────────────────────

@app.route("/api/scenes/<int:idx>/reparse", methods=["POST"])
def reparse_heading(idx):
    """Re-parse the scene heading to extract int_ext, location, time_of_day."""
    state = AppState()
    if not state.scene_list or idx < 0 or idx >= len(state.scene_list):
        return jsonify({"error": "无效的场次索引"}), 400

    scene = state.scene_list[idx]

    # Allow passing a custom heading text for parsing, or use stored heading
    data = request.get_json(silent=True) or {}
    heading_text = (data.get("heading") or "").strip() or scene.heading

    fields = parse_heading_fields(heading_text)

    state.push_undo()
    scene.heading = heading_text
    scene.int_ext = fields["int_ext"]
    scene.location = fields["location"]
    scene.time_of_day = fields["time_of_day"]
    scene.is_manually_adjusted = True
    state.status_message = f"已重新解析场次 {idx + 1} 标题"
    return jsonify({"fields": fields, **state.to_dict()})


# ── Rule-based Detect ────────────────────────────────────────────

@app.route("/api/detect", methods=["POST"])
def detect():
    state = AppState()
    if not state.parse_result:
        return jsonify({"error": "请先上传剧本文件"}), 400

    state.push_undo()

    detector = SceneDetector()
    state.scene_list = detector.detect(state.parse_result)
    state.status_message = f"自动识别完成: {len(state.scene_list)} 个场次"
    return jsonify(state.to_dict())


# ── LLM Detect (async) ──────────────────────────────────────────

@app.route("/api/llm/detect", methods=["POST"])
def llm_detect():
    state = AppState()
    if not state.parse_result:
        return jsonify({"error": "请先上传剧本文件"}), 400
    if not state.llm_config.provider:
        return jsonify({"error": "请先配置 LLM 设置"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有 LLM 任务运行中"}), 409

    state.push_undo()

    state.llm_task.task_type = "detect"
    state.llm_task.status = "running"
    state.llm_task.error = ""
    state.llm_task.result = None
    state.status_message = "LLM 场次识别中..."

    def _run():
        try:
            llm = _create_llm(state.llm_config)
            llm_detector = LLMSceneDetector(llm)
            result = llm_detector.detect(state.parse_result)

            # ── 保留已手动调整场次的基本内容 ──────────────────────
            # 将 LLM 新识别的场次与旧场次按 start_line 就近匹配；
            # 若旧场次被手动调整过，则将其基本信息写回到新场次，
            # 确保 LLM 辅助识别不覆盖用户手动校正的内容。
            if state.scene_list:
                old_list = list(state.scene_list)
                for new_s in result:
                    closest = min(
                        old_list,
                        key=lambda s: abs(s.start_line - new_s.start_line),
                    )
                    if (
                        abs(closest.start_line - new_s.start_line) <= 10
                        and closest.is_manually_adjusted
                    ):
                        # Only copy non-empty values so LLM-extracted fields
                        # are not overwritten by empty strings from rule detection
                        if closest.heading:     new_s.heading     = closest.heading
                        if closest.location:    new_s.location    = closest.location
                        if closest.int_ext:     new_s.int_ext     = closest.int_ext
                        if closest.time_of_day: new_s.time_of_day = closest.time_of_day
                        if closest.scene_type:  new_s.scene_type  = closest.scene_type
                        if closest.summary:     new_s.summary     = closest.summary
                        new_s.is_manually_adjusted = True

            state.scene_list = result
            state.llm_task.status = "done"
            state.llm_task.result = True
            state.status_message = f"LLM 识别完成: {len(result)} 个场次"
        except Exception as e:
            state.llm_task.status = "error"
            state.llm_task.error = str(e)
            state.status_message = f"LLM 错误: {e}"

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "running", "task_type": "detect"})


# ── LLM PDF Clean (async) ────────────────────────────────────────

@app.route("/api/llm/pdf-clean", methods=["POST"])
def llm_pdf_clean():
    state = AppState()
    if not state.parse_result:
        return jsonify({"error": "请先上传剧本文件"}), 400
    if not state.filename or not state.filename.lower().endswith(".pdf"):
        return jsonify({"error": "仅支持 PDF 文件"}), 400
    if not state.llm_config.provider:
        return jsonify({"error": "请先配置 LLM 设置"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有 LLM 任务运行中"}), 409

    state.push_undo()

    state.llm_task.task_type = "pdf_clean"
    state.llm_task.status = "running"
    state.llm_task.error = ""
    state.llm_task.result = None
    state.status_message = "LLM 分析 PDF 场次结构..."

    def _run():
        try:
            llm = _create_llm(state.llm_config)
            lines = state.parse_result.lines

            # Process in chunks of 80 lines to keep context manageable
            chunk_size = 80
            all_headings: list[dict] = []
            for start in range(0, len(lines), chunk_size):
                chunk = lines[start:start + chunk_size]
                headings = llm.identify_scene_headings(chunk, line_offset=start)
                all_headings.extend(headings)

            if not all_headings:
                state.llm_task.status = "error"
                state.llm_task.error = "LLM 未能识别出任何场次标题"
                state.status_message = "PDF 识别失败: 未找到场次"
                return

            # Store structured info in line_metadata so SceneDetector can use it
            line_metadata: dict[int, dict] = {}
            for h in all_headings:
                idx = h.get("line_index", -1)
                if isinstance(idx, int) and 0 <= idx < len(lines):
                    line_metadata[idx] = {
                        "type": "Scene Heading",
                        "int_ext": h.get("int_ext", ""),
                        "location": h.get("location", ""),
                        "time_of_day": h.get("time_of_day", ""),
                    }

            state.parse_result.line_metadata = line_metadata

            # Re-detect scenes using the LLM-identified metadata
            result = SceneDetector().detect(state.parse_result)
            state.scene_list = result
            state.llm_task.status = "done"
            state.llm_task.result = True
            state.status_message = f"PDF LLM 识别完成: {len(result)} 个场次"
        except Exception as e:
            state.llm_task.status = "error"
            state.llm_task.error = str(e)
            state.status_message = f"PDF 识别失败: {e}"

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "running", "task_type": "pdf_clean"})


# ── LLM Summarize (async) ───────────────────────────────────────

@app.route("/api/llm/summarize/<int:idx>", methods=["POST"])
def llm_summarize(idx):
    state = AppState()
    if not state.scene_list or idx < 0 or idx >= len(state.scene_list):
        return jsonify({"error": "无效的场次索引"}), 400
    if not state.llm_config.provider:
        return jsonify({"error": "请先配置 LLM 设置"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有 LLM 任务运行中"}), 409

    state.push_undo()

    state.llm_task.task_type = "summarize"
    state.llm_task.status = "running"
    state.llm_task.error = ""
    state.llm_task.result = None
    state.llm_task.scene_index = idx
    state.status_message = f"正在为场次 {idx + 1} 生成摘要..."

    def _run():
        try:
            llm = _create_llm(state.llm_config)
            llm_detector = LLMSceneDetector(llm)
            scene = state.scene_list[idx]
            summary = llm_detector.summarize_single_scene(scene)
            state.llm_task.status = "done"
            state.llm_task.result = summary
            state.status_message = f"场次 {idx + 1} 摘要已生成"
        except Exception as e:
            state.llm_task.status = "error"
            state.llm_task.error = str(e)
            state.status_message = f"摘要生成失败: {e}"

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "running", "task_type": "summarize", "scene_index": idx})


# ── LLM Task Status ─────────────────────────────────────────────

@app.route("/api/llm/status")
def llm_status():
    state = AppState()
    resp = {
        "task_type": state.llm_task.task_type,
        "status": state.llm_task.status,
        "error": state.llm_task.error,
        "scene_index": state.llm_task.scene_index,
    }

    # When done, include full state update so frontend can refresh
    if state.llm_task.status in ("done", "error"):
        resp["state"] = state.to_dict()
        # Reset task after client reads it
        if state.llm_task.status == "done":
            state.llm_task.status = "idle"

    return jsonify(resp)


# ── LLM Config ───────────────────────────────────────────────────

@app.route("/api/llm/config", methods=["GET", "PUT"])
def llm_config():
    state = AppState()

    if request.method == "GET":
        cfg = state.llm_config
        return jsonify({
            "provider": cfg.provider,
            "model_name": cfg.model_name,
            "api_key": "***" if cfg.api_key else "",
            "base_url": cfg.base_url,
            "temperature": cfg.temperature,
            "max_tokens": cfg.max_tokens,
        })

    data = request.get_json()
    # Preserve existing api_key if user left the field empty
    new_key = data.get("api_key", "").strip()
    api_key = new_key if new_key else state.llm_config.api_key
    state.llm_config = LLMConfig(
        provider=data.get("provider", ""),
        model_name=data.get("model_name", ""),
        api_key=api_key,
        base_url=data.get("base_url", ""),
        temperature=data.get("temperature", 0.3),
        max_tokens=data.get("max_tokens", 4096),
    )
    _save_llm_config(state.llm_config)
    state.status_message = f"LLM 设置已更新: {state.llm_config.provider}"
    return jsonify({"ok": True})


# ── LLM Test Connection ─────────────────────────────────────────

@app.route("/api/llm/test", methods=["POST"])
def llm_test():
    state = AppState()
    data = request.get_json()
    # Use existing api_key if none provided
    new_key = data.get("api_key", "").strip()
    api_key = new_key if new_key else state.llm_config.api_key
    config = LLMConfig(
        provider=data.get("provider", ""),
        model_name=data.get("model_name", ""),
        api_key=api_key,
        base_url=data.get("base_url", ""),
        temperature=data.get("temperature", 0.3),
        max_tokens=data.get("max_tokens", 4096),
    )
    try:
        llm = _create_llm(config)
        ok = llm.test_connection()
        return jsonify({"ok": ok})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── Line Delete ──────────────────────────────────────────────────

@app.route("/api/lines/delete", methods=["POST"])
def delete_lines():
    """Delete one or more raw script lines, adjusting scene boundaries."""
    state = AppState()
    if not state.parse_result or not state.scene_list:
        return jsonify({"error": "请先上传并识别剧本"}), 400

    data = request.get_json()
    line_indices = data.get("line_indices", [])
    if not isinstance(line_indices, list) or not line_indices:
        return jsonify({"error": "缺少 line_indices"}), 400

    # Validate indices
    total = len(state.parse_result.lines)
    valid = [i for i in line_indices if isinstance(i, int) and 0 <= i < total]
    if not valid:
        return jsonify({"error": "无效的行号"}), 400

    state.push_undo()

    new_lines = state.scene_list.delete_lines(valid, state.parse_result.lines)
    state.parse_result = ParseResult(
        lines=new_lines,
        metadata=state.parse_result.metadata,
        line_metadata={},   # position-based metadata is now stale
    )

    count = len(valid)
    state.status_message = f"已删除 {count} 行"
    return jsonify(state.to_dict())


# ── Line Replace (full content edit) ────────────────────────────

@app.route("/api/lines/replace", methods=["POST"])
def lines_replace():
    """Replace all script lines with user-edited content."""
    state = AppState()
    if not state.parse_result:
        return jsonify({"error": "请先上传剧本文件"}), 400

    data = request.get_json() or {}
    new_lines = data.get("lines", [])
    if not isinstance(new_lines, list):
        return jsonify({"error": "lines 必须为数组"}), 400

    state.push_undo()

    state.parse_result = ParseResult(
        lines=[str(l) for l in new_lines],
        metadata=state.parse_result.metadata,
        line_metadata={},  # stale after full edit
    )

    # Re-run rule-based scene detection to keep scene list in sync
    result = SceneDetector().detect(state.parse_result)
    state.scene_list = result

    state.status_message = f"剧本内容已更新，共 {len(new_lines)} 行"
    return jsonify(state.to_dict())


# ── Content Lock Toggle ──────────────────────────────────────────

@app.route("/api/lines/toggle-lock", methods=["POST"])
def lines_toggle_lock():
    """Toggle content_locked flag."""
    state = AppState()
    state.content_locked = not state.content_locked
    state.status_message = "内容已锁定" if state.content_locked else "内容已解锁，可自由编辑"
    return jsonify(state.to_dict())


# ── Calibration ──────────────────────────────────────────────────

@app.route("/api/calibrate/insert_break", methods=["POST"])
def insert_break():
    state = AppState()
    if not state.scene_list or not state.parse_result:
        return jsonify({"error": "请先上传并识别剧本"}), 400

    data = request.get_json()
    line_index = data.get("line_index")
    if line_index is None:
        return jsonify({"error": "缺少 line_index"}), 400

    state.push_undo()

    state.scene_list.insert_break(line_index, state.parse_result.lines)
    state.status_message = f"已在第 {line_index + 1} 行插入场次分隔"
    return jsonify(state.to_dict())


@app.route("/api/calibrate/delete_break", methods=["POST"])
def delete_break():
    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "请先上传并识别剧本"}), 400

    data = request.get_json()
    line_index = data.get("line_index")
    if line_index is None:
        return jsonify({"error": "缺少 line_index"}), 400

    state.push_undo()

    n = len(state.scene_list)
    for i, scene in enumerate(state.scene_list):
        if scene.start_line == line_index and i > 0:
            # Build reindex mapping: scene i is removed, scenes after i shift by -1
            old_to_new = {j: j for j in range(i)}
            old_to_new.update({j: j - 1 for j in range(i + 1, n)})
            state.scene_list.remove_scene(i)
            state.reindex_entities(old_to_new)
            state.status_message = f"已删除第 {line_index + 1} 行的场次分隔"
            return jsonify(state.to_dict())

    return jsonify({"error": "该行不是场次起始行"}), 400


@app.route("/api/calibrate/merge", methods=["POST"])
def merge_scenes():
    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "请先上传并识别剧本"}), 400

    data = request.get_json()
    idx1 = data.get("index1")
    idx2 = data.get("index2")
    if idx1 is None or idx2 is None:
        return jsonify({"error": "缺少 index1/index2"}), 400

    state.push_undo()

    lo, hi = min(idx1, idx2), max(idx1, idx2)
    n = len(state.scene_list)
    state.scene_list.merge_scenes(idx1, idx2)
    # hi is merged into lo; scenes after hi shift by -1
    old_to_new = {j: j for j in range(hi)}
    old_to_new.update({j: j - 1 for j in range(hi + 1, n)})
    state.reindex_entities(old_to_new)
    state.status_message = f"已合并场次 {idx1 + 1} 和 {idx2 + 1}"
    return jsonify(state.to_dict())


# ── Entity Extraction ────────────────────────────────────────────

@app.route("/api/extract/<int:idx>", methods=["POST"])
def extract_single(idx):
    """Extract entities from a single scene (rules, or LLM if use_llm=true in body)."""
    state = AppState()
    if not state.scene_list or idx < 0 or idx >= len(state.scene_list):
        return jsonify({"error": "无效的场次索引"}), 400

    scene = state.scene_list[idx]
    body = request.get_json(silent=True) or {}
    use_llm = bool(body.get("use_llm", False))

    if use_llm:
        if not state.llm_config.provider:
            return jsonify({"error": "请先在「LLM 设置」中配置 API 密钥后再使用 LLM 辅助提取"}), 400
        if state.llm_task.status == "running":
            return jsonify({"error": "已有 LLM 任务运行中"}), 409

        state.llm_task.task_type = "extract"
        state.llm_task.status = "running"
        state.llm_task.error = ""
        state.llm_task.result = None
        state.llm_task.scene_index = idx
        state.status_message = f"LLM 正在提取场次 {idx + 1} 的实体..."

        # 预先构建场景类型键列表，供 LLM 分类时使用
        try:
            _pre_params = _get_duration_params(state)
            _pre_engine = SceneDurationEngine(_pre_params)
            _scene_type_keys = (
                list(_pre_params.table_a.keys()) +
                list(_pre_params.custom_scene_types.keys())
            )
        except Exception:
            _pre_engine = None
            _scene_type_keys = []

        def _run():
            try:
                llm = _create_llm(state.llm_config)
                llm_detector = LLMSceneDetector(llm)
                result = llm_detector.extract_entities(scene, scene_type_keys=_scene_type_keys)
                # ── LLM 提取后验证 scene_type_key，无效则回退规则分类 ──
                try:
                    llm_key = result.get("scene_type_key", "")
                    if llm_key and llm_key in _scene_type_keys:
                        # LLM 返回了有效的类型键，构造分类结果
                        result["classification"] = {
                            "a_key": llm_key,
                            "confidence": "high",
                            "method": "llm",
                            "matched_keywords": [],
                            "detail": "LLM 识别场景类型",
                        }
                    else:
                        # 回退到规则分类
                        engine = _pre_engine or SceneDurationEngine(_get_duration_params(state))
                        _cls = engine.classify_scene(scene, result)
                        result["scene_type_key"] = _cls["a_key"]
                        result["classification"] = _cls
                except Exception:
                    pass
                state.entities[idx] = result
                state.rebuild_global_entities()
                state.llm_task.status = "done"
                state.llm_task.result = result
                state.status_message = f"场次 {idx + 1} 实体提取完成（LLM）"
            except Exception as e:
                state.llm_task.status = "error"
                state.llm_task.error = str(e)
                state.status_message = f"实体提取失败: {e}"

        threading.Thread(target=_run, daemon=True).start()
        return jsonify({"status": "running", "task_type": "extract", "scene_index": idx})
    else:
        # Rules-only extraction (synchronous)
        result = rule_extract_entities(scene)
        # 同步运行规则分类并附加到结果
        try:
            _params = _get_duration_params(state)
            _engine = SceneDurationEngine(_params)
            _cls = _engine.classify_scene(scene, result)
            result["scene_type_key"] = _cls["a_key"]
            result["classification"] = _cls
        except Exception:
            pass
        state.entities[idx] = result
        state.rebuild_global_entities()
        state.status_message = f"场次 {idx + 1} 实体提取完成"
        return jsonify(state.to_dict())


@app.route("/api/extract/all", methods=["POST"])
def extract_all():
    """Extract entities from all scenes (async, optionally with LLM)."""
    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "请先上传并识别剧本"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有 LLM 任务运行中"}), 409

    body = request.get_json(silent=True) or {}
    use_llm = bool(body.get("use_llm", False))

    if use_llm and not state.llm_config.provider:
        return jsonify({"error": "请先在「LLM 设置」中配置 API 密钥后再使用 LLM 辅助提取"}), 400

    state.llm_task.task_type = "extract"
    state.llm_task.status = "running"
    state.llm_task.error = ""
    state.llm_task.result = None
    state.llm_task.scene_index = -1
    label = "LLM" if use_llm else "规则"
    state.status_message = f"正在提取所有场次的实体（{label}）..."

    def _run():
        try:
            llm_detector = None
            if use_llm:
                llm = _create_llm(state.llm_config)
                llm_detector = LLMSceneDetector(llm)

            _params = _get_duration_params(state)
            _engine = SceneDurationEngine(_params)
            _type_keys = (
                list(_params.table_a.keys()) +
                list(_params.custom_scene_types.keys())
            )
            for i, scene in enumerate(state.scene_list):
                if llm_detector:
                    result = llm_detector.extract_entities(scene, scene_type_keys=_type_keys)
                else:
                    result = rule_extract_entities(scene)
                try:
                    llm_key = result.get("scene_type_key", "")
                    if llm_key and llm_key in _type_keys:
                        result["classification"] = {
                            "a_key": llm_key,
                            "confidence": "high",
                            "method": "llm",
                            "matched_keywords": [],
                            "detail": "LLM 识别场景类型",
                        }
                    else:
                        _cls = _engine.classify_scene(scene, result)
                        result["scene_type_key"] = _cls["a_key"]
                        result["classification"] = _cls
                except Exception:
                    pass
                state.entities[i] = result

            state.rebuild_global_entities()
            state.llm_task.status = "done"
            state.llm_task.result = True
            state.status_message = f"全部 {len(state.scene_list)} 个场次实体提取完成（{label}）"
        except Exception as e:
            state.llm_task.status = "error"
            state.llm_task.error = str(e)
            state.status_message = f"实体提取失败: {e}"

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "running", "task_type": "extract"})


@app.route("/api/entities")
def get_entities():
    """Get all extraction results."""
    state = AppState()
    return jsonify({
        "entities": state.entities,
        "global_entities": state.global_entities,
    })


@app.route("/api/entities/<int:idx>", methods=["GET", "PUT"])
def entity_detail(idx):
    state = AppState()
    if request.method == "GET":
        return jsonify(state.entities.get(idx, {"characters": [], "props": [], "scene_type": ""}))

    # PUT: update entities for a specific scene
    if not state.scene_list or idx < 0 or idx >= len(state.scene_list):
        return jsonify({"error": "无效的场次索引"}), 400

    data = request.get_json()
    state.push_undo()
    current = dict(state.entities.get(idx, {"characters": [], "props": [], "scene_type": ""}))
    if "characters" in data:
        current["characters"] = [x for x in data["characters"] if x]
    if "props" in data:
        current["props"] = [x for x in data["props"] if x]
    if "scene_type" in data:
        current["scene_type"] = data["scene_type"]
    if "scene_type_key" in data:
        current["scene_type_key"] = data["scene_type_key"]
    if "classification" in data:
        current["classification"] = data["classification"]
    if "duration_override_hours" in data:
        val = data["duration_override_hours"]
        current["duration_override_hours"] = float(val) if val not in (None, "") else None
    state.entities[idx] = current
    state.rebuild_global_entities()
    state.status_message = f"场次 {idx + 1} 实体已更新"
    return jsonify(state.to_dict())


@app.route("/api/entities/global")
def get_global_entities():
    """Get global entity summary."""
    state = AppState()
    return jsonify(state.global_entities)


@app.route("/api/entities/rename", methods=["PUT"])
def rename_entity():
    """Rename a character or prop globally across all scenes."""
    state = AppState()
    data = request.get_json()
    old_name = (data.get("old_name") or "").strip()
    new_name = (data.get("new_name") or "").strip()
    entity_type = data.get("type", "character")

    if not old_name or not new_name:
        return jsonify({"error": "缺少参数"}), 400
    if old_name == new_name:
        return jsonify(state.to_dict())

    state.push_undo()
    state.rename_entity(old_name, new_name, entity_type)
    label = "人物" if entity_type == "character" else "道具"
    state.status_message = f"已将{label}「{old_name}」重命名为「{new_name}」"
    return jsonify(state.to_dict())


@app.route("/api/entities/remove", methods=["DELETE"])
def remove_entity():
    """Remove a character or prop globally from all scenes."""
    state = AppState()
    data = request.get_json()
    name = (data.get("name") or "").strip()
    entity_type = data.get("type", "character")

    if not name:
        return jsonify({"error": "缺少参数"}), 400

    state.push_undo()
    state.remove_entity(name, entity_type)
    label = "人物" if entity_type == "character" else "道具"
    state.status_message = f"已删除{label}「{name}」"
    return jsonify(state.to_dict())


@app.route("/api/entities/merge", methods=["POST"])
def merge_entities():
    """Merge multiple entity names into one target name."""
    state = AppState()
    data = request.get_json()
    target_name = (data.get("target_name") or "").strip()
    source_names = [s.strip() for s in data.get("source_names", []) if s and s.strip()]
    entity_type = data.get("type", "character")

    if not target_name or not source_names:
        return jsonify({"error": "缺少参数"}), 400

    state.push_undo()
    for src in source_names:
        if src != target_name:
            state.rename_entity(src, target_name, entity_type)
    label = "人物" if entity_type == "character" else "道具"
    state.status_message = f"已合并 {len(source_names)} 个{label}名称为「{target_name}」"
    return jsonify(state.to_dict())


# ── LLM Character Analysis ──────────────────────────────────────

@app.route("/api/llm/analyze_character", methods=["POST"])
def analyze_character():
    """Analyze a single character across relevant scenes."""
    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "请先上传并识别剧本"}), 400
    if not state.llm_config.provider:
        return jsonify({"error": "请先配置 LLM 设置"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有 LLM 任务运行中"}), 409

    data = request.get_json()
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "缺少角色名"}), 400

    state.llm_task.task_type = "analyze_character"
    state.llm_task.status = "running"
    state.llm_task.error = ""
    state.llm_task.result = None
    state.status_message = f"正在分析角色: {name}..."

    def _run():
        try:
            llm = _create_llm(state.llm_config)
            llm_detector = LLMSceneDetector(llm)

            # Find scenes where this character appears
            relevant_scenes = []
            for idx, ent in state.entities.items():
                idx_int = int(idx) if isinstance(idx, str) else idx
                if name in ent.get("characters", []):
                    if 0 <= idx_int < len(state.scene_list):
                        relevant_scenes.append(state.scene_list[idx_int])

            if not relevant_scenes:
                # Fallback: search all scenes for character name
                relevant_scenes = [
                    s for s in state.scene_list
                    if name in s.content
                ]

            analysis = llm_detector.analyze_character(name, relevant_scenes)
            state.character_analyses[name] = analysis
            state.llm_task.status = "done"
            state.llm_task.result = analysis
            state.status_message = f"角色 {name} 分析完成"
        except Exception as e:
            state.llm_task.status = "error"
            state.llm_task.error = str(e)
            state.status_message = f"角色分析失败: {e}"

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "running", "task_type": "analyze_character"})


@app.route("/api/llm/analyze_all", methods=["POST"])
def analyze_all_characters():
    """Global character relationship analysis."""
    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "请先上传并识别剧本"}), 400
    if not state.llm_config.provider:
        return jsonify({"error": "请先配置 LLM 设置"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有 LLM 任务运行中"}), 409

    # Collect character names from global entities
    character_names = [c["name"] for c in state.global_entities.get("characters", [])]
    if not character_names:
        return jsonify({"error": "请先执行实体提取"}), 400

    state.llm_task.task_type = "analyze_all"
    state.llm_task.status = "running"
    state.llm_task.error = ""
    state.llm_task.result = None
    state.status_message = "正在进行全局人物分析..."

    def _run():
        try:
            llm = _create_llm(state.llm_config)
            llm_detector = LLMSceneDetector(llm)
            analysis = llm_detector.analyze_characters_global(state.scene_list, character_names)
            state.global_analysis = analysis
            state.llm_task.status = "done"
            state.llm_task.result = analysis
            state.status_message = "全局人物分析完成"
        except Exception as e:
            state.llm_task.status = "error"
            state.llm_task.error = str(e)
            state.status_message = f"全局分析失败: {e}"

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "running", "task_type": "analyze_all"})


# ── Episode Synopsis ──────────────────────────────────────────────

@app.route("/api/episodes/current/synopsis", methods=["PUT"])
def save_episode_synopsis():
    """Save the episode synopsis (user-written or LLM-generated)."""
    state = AppState()
    data = request.get_json() or {}
    state.episode_synopsis = data.get("synopsis", "")
    state.save_to_episode()
    return jsonify({"ok": True})


@app.route("/api/episodes/current/synopsis/generate", methods=["POST"])
def generate_episode_synopsis():
    """Generate episode synopsis via LLM (async)."""
    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "请先上传并识别剧本"}), 400
    if not state.llm_config.model_name or not state.llm_config.api_key:
        return jsonify({"error": "请先配置 LLM 设置"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有 LLM 任务运行中"}), 409

    state.llm_task.task_type = "synopsis"
    state.llm_task.status = "running"
    state.llm_task.error = ""
    state.status_message = "正在生成集数梗概..."

    # Build scene list for the prompt
    scenes_text = []
    for s in state.scene_list:
        line = f"场次 {s.scene_number}: {s.heading}"
        if s.summary:
            line += f"\n  摘要: {s.summary}"
        scenes_text.append(line)
    content = "\n".join(scenes_text)

    def _run():
        try:
            llm = _create_llm(state.llm_config)
            prompt = (
                "请根据以下剧本场次信息，为这集剧本写一段简洁的梗概（约100-200字），"
                "概括主要情节、人物关系和核心冲突。\n\n"
                f"{content}\n\n"
                "请直接返回梗概文字，不要包含任何额外格式标注。"
            )
            result = llm.complete(prompt)
            state.episode_synopsis = result.strip()
            state.save_to_episode()
            state.llm_task.status = "done"
            state.status_message = "梗概生成完成"
        except Exception as e:
            state.llm_task.status = "error"
            state.llm_task.error = str(e)
            state.status_message = f"梗概生成失败: {e}"

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "running", "task_type": "synopsis"})


# ── Location Rename (across current episode) ─────────────────────

@app.route("/api/scenes/location/rename", methods=["PUT"])
def rename_location():
    """Rename a location across all scenes in the current episode."""
    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "请先上传并识别剧本"}), 400

    data = request.get_json()
    old_name = (data.get("old_name") or "").strip()
    new_name = (data.get("new_name") or "").strip()

    if not old_name:
        return jsonify({"error": "缺少参数"}), 400
    if old_name == new_name:
        return jsonify(state.to_dict())

    state.push_undo()
    count = 0
    for scene in state.scene_list:
        if (scene.location or "").strip() == old_name:
            scene.location = new_name
            scene.is_manually_adjusted = True
            count += 1

    state.status_message = f"已将地点「{old_name}」重命名为「{new_name}」（影响 {count} 个场次）"
    return jsonify(state.to_dict())


# ── Scene Delete (Batch) ─────────────────────────────────────────

@app.route("/api/scenes/delete", methods=["POST"])
def delete_scenes():
    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "请先上传并识别剧本"}), 400

    data = request.get_json()
    indices = data.get("indices", [])

    if not indices:
        return jsonify({"error": "未指定场次索引"}), 400

    valid = [i for i in indices if 0 <= i < len(state.scene_list)]
    if not valid:
        return jsonify({"error": "无效的场次索引"}), 400
    if len(valid) >= len(state.scene_list):
        return jsonify({"error": "不允许删除全部场次"}), 400

    state.push_undo()
    old_to_new = state.scene_list.batch_remove_scenes(valid)
    state.reindex_entities(old_to_new)
    state.status_message = f"已删除 {len(valid)} 个场次"
    return jsonify(state.to_dict())


# ── Project Management ────────────────────────────────────────────

@app.route("/api/projects", methods=["GET"])
def list_projects():
    return jsonify({"projects": Project.list_projects()})


@app.route("/api/projects", methods=["POST"])
def create_project():
    data = request.get_json()
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "请输入项目名称"}), 400

    state = AppState()
    project = Project.create_new(name)

    # If a script is already loaded, preserve it as the first episode
    if state.parse_result and state.filename:
        episode_name = os.path.splitext(state.filename)[0]
        episode = project.add_episode(episode_name, state.filename)
        state.project = project
        state.current_episode_id = episode.id
        project.meta.active_episode_id = episode.id
        state.save_to_episode()
        state.status_message = f"已创建项目: {name}，当前剧本已保存为第 1 集「{episode_name}」"
    else:
        state.reset()
        state.project = project
        state.status_message = f"已创建项目: {name}"

    _save_last_project(project.meta.id)
    return jsonify(state.to_dict())


@app.route("/api/projects/<project_id>/open", methods=["POST"])
def open_project(project_id):
    try:
        project = Project.load(project_id)
    except FileNotFoundError:
        return jsonify({"error": "项目不存在"}), 404

    state = AppState()
    state.reset()
    state.project = project

    # Load active episode if exists
    if project.meta.active_episode_id and project.meta.episode_order:
        try:
            state.load_from_episode(project.meta.active_episode_id)
        except Exception as e:
            state.status_message = f"已打开项目: {project.meta.name}"
    else:
        state.status_message = f"已打开项目: {project.meta.name} (无集数)"

    _save_last_project(project.meta.id)
    return jsonify(state.to_dict())


@app.route("/api/projects/current/save", methods=["POST"])
def save_project():
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    state.save_to_episode()
    state.project._save_meta()
    state.status_message = "项目已保存"
    return jsonify({"ok": True, "status": state.status_message})


@app.route("/api/projects/current", methods=["DELETE"])
def delete_current_project():
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    state.project.delete_project()
    state.reset()
    _clear_last_project()
    state.status_message = "项目已删除"
    return jsonify(state.to_dict())


# ── Episode Management ────────────────────────────────────────────

@app.route("/api/episodes", methods=["GET"])
def list_episodes():
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400
    return jsonify({
        "episodes": state.project.get_episodes_info(),
        "active_episode_id": state.project.meta.active_episode_id,
    })


@app.route("/api/episodes/add", methods=["POST"])
def add_episode():
    state = AppState()
    if not state.project:
        return jsonify({"error": "请先创建或打开项目"}), 400

    if "file" not in request.files:
        return jsonify({"error": "没有上传文件"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "文件名为空"}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    parser_cls = PARSERS.get(ext)
    if parser_cls is None:
        return jsonify({"error": f"不支持的文件格式: {ext}"}), 400

    episode_name = request.form.get("name", "").strip() or os.path.splitext(file.filename)[0]

    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        file.save(tmp)
        tmp_path = tmp.name

    try:
        # Save current episode before switching
        state.save_to_episode()

        # Parse the new file
        parser = parser_cls()
        parse_result = parser.parse(tmp_path)

        # Create episode in project
        episode = state.project.add_episode(episode_name, file.filename)

        # Load new episode into state
        state.parse_result = parse_result
        state.filename = file.filename
        state.scene_list = None
        state.entities.clear()
        state.global_entities = {"characters": [], "props": []}
        state.character_analyses.clear()
        state.global_analysis = ""
        state._undo_stack.clear()
        state._redo_stack.clear()
        state.current_episode_id = episode.id
        state.project.meta.active_episode_id = episode.id

        # Auto-detect scenes
        detector = SceneDetector()
        state.scene_list = detector.detect(state.parse_result)
        state.push_undo()

        state.status_message = (
            f"已导入: {file.filename} ({len(parse_result.lines)} 行), "
            f"自动识别 {len(state.scene_list)} 个场次"
        )

        # Save new episode state
        state.save_to_episode()

        return jsonify(state.to_dict())

    except Exception as e:
        return jsonify({"error": f"解析文件失败: {e}"}), 500
    finally:
        os.unlink(tmp_path)


@app.route("/api/episodes/<episode_id>/switch", methods=["POST"])
def switch_episode(episode_id):
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    if episode_id not in state.project.meta.episode_order:
        return jsonify({"error": "集数不存在"}), 404

    if episode_id == state.current_episode_id:
        return jsonify(state.to_dict())

    # Save current episode
    state.save_to_episode()

    # Load target episode
    try:
        state.load_from_episode(episode_id)
        state.project._save_meta()
    except Exception as e:
        return jsonify({"error": f"切换失败: {e}"}), 500

    return jsonify(state.to_dict())


@app.route("/api/episodes/<episode_id>/rename", methods=["PUT"])
def rename_episode(episode_id):
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    data = request.get_json()
    new_name = (data.get("name") or "").strip()
    if not new_name:
        return jsonify({"error": "名称不能为空"}), 400

    try:
        state.project.rename_episode(episode_id, new_name)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    state.status_message = f"已重命名集数为: {new_name}"
    return jsonify(state.to_dict())


@app.route("/api/episodes/<episode_id>", methods=["DELETE"])
def delete_episode(episode_id):
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    try:
        was_active = (episode_id == state.current_episode_id)
        state.project.remove_episode(episode_id)

        if was_active:
            new_active = state.project.meta.active_episode_id
            state.load_from_episode(new_active)
            state.project._save_meta()
        else:
            state.status_message = "集数已删除"

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"删除失败: {e}"}), 500

    return jsonify(state.to_dict())


@app.route("/api/episodes/reorder", methods=["POST"])
def reorder_episodes():
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    data = request.get_json()
    new_order = data.get("order", [])

    try:
        state.project.reorder_episodes(new_order)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    state.status_message = "集数顺序已更新"
    return jsonify(state.to_dict())


# ── Project Summary (cross-episode overview) ─────────────────────

@app.route("/api/projects/summary")
def project_summary():
    """Return aggregated scene/character/prop data across all episodes."""
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    # Save current episode so its data is up-to-date on disk
    state.save_to_episode()

    episodes_info = []
    all_chars: dict = {}   # name -> {episodes: set, total_scenes: int}
    all_props: dict = {}   # name -> {episodes: set, total_scenes: int}
    total_scenes = 0

    for ep_id in state.project.meta.episode_order:
        try:
            ep = state.project.load_episode(ep_id)
        except Exception:
            continue

        scenes = ep.scenes or []
        entities = ep.entities or {}
        ep_chars: set = set()
        ep_props: set = set()

        for ent in entities.values():
            for ch in ent.get("characters", []):
                ep_chars.add(ch)
                rec = all_chars.setdefault(ch, {"episodes": set(), "total_scenes": 0})
                rec["episodes"].add(ep.name)
                rec["total_scenes"] += 1
            for pr in ent.get("props", []):
                ep_props.add(pr)
                rec = all_props.setdefault(pr, {"episodes": set(), "total_scenes": 0})
                rec["episodes"].add(ep.name)
                rec["total_scenes"] += 1

        sc = len(scenes)
        total_scenes += sc
        episodes_info.append({
            "id": ep_id,
            "name": ep.name,
            "filename": ep.filename,
            "scene_count": sc,
            "character_count": len(ep_chars),
            "prop_count": len(ep_props),
            "is_active": ep_id == state.current_episode_id,
        })

    return jsonify({
        "project_name": state.project.meta.name,
        "total_scenes": total_scenes,
        "episodes": episodes_info,
        "all_characters": sorted(
            [{"name": k, "episode_count": len(v["episodes"]),
              "episodes": sorted(v["episodes"]), "total_scenes": v["total_scenes"]}
             for k, v in all_chars.items()],
            key=lambda x: -x["total_scenes"]),
        "all_props": sorted(
            [{"name": k, "episode_count": len(v["episodes"]),
              "episodes": sorted(v["episodes"]), "total_scenes": v["total_scenes"]}
             for k, v in all_props.items()],
            key=lambda x: -x["total_scenes"]),
    })


# ── Detailed Overview (cross-episode) ────────────────────────────

@app.route("/api/overview/full")
def overview_full():
    """Return detailed cross-episode overview: characters, locations, props with scene lists."""
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    state.save_to_episode()

    chars: dict = {}      # name -> {ep_name: [scene_info, ...]}
    props: dict = {}      # name -> {ep_name: [scene_info, ...]}
    locations: dict = {}  # loc  -> {ep_name: [scene_info, ...]}
    episodes_info = []
    total_scenes = 0

    for ep_id in state.project.meta.episode_order:
        try:
            ep = state.project.load_episode(ep_id)
        except Exception:
            continue

        scenes = ep.scenes or []
        entities = ep.entities or {}
        total_scenes += len(scenes)
        episodes_info.append({"id": ep_id, "name": ep.name, "scene_count": len(scenes)})

        for i, sd in enumerate(scenes):
            e = entities.get(i) or entities.get(str(i)) or {}
            si = {
                "scene_number": sd.get("scene_number", i + 1),
                "heading": sd.get("heading", ""),
                "location": sd.get("location", ""),
                "time_of_day": sd.get("time_of_day", ""),
                "int_ext": sd.get("int_ext", ""),
                "episode_name": ep.name,
                "episode_id": ep_id,
                "characters": e.get("characters", []),
                "props": e.get("props", []),
            }

            # Group by character
            for ch in e.get("characters", []):
                chars.setdefault(ch, {}).setdefault(ep.name, []).append(si)

            # Group by prop
            for pr in e.get("props", []):
                props.setdefault(pr, {}).setdefault(ep.name, []).append(si)

            # Group by location
            loc = (sd.get("location") or "").strip() or "（未知地点）"
            locations.setdefault(loc, {}).setdefault(ep.name, []).append(si)

    def _fmt(name_map):
        result = []
        for name, ep_map in name_map.items():
            total = sum(len(v) for v in ep_map.values())
            result.append({
                "name": name,
                "total_scenes": total,
                "episodes": [{"episode_name": epn, "scenes": slist}
                              for epn, slist in ep_map.items()],
            })
        result.sort(key=lambda x: -x["total_scenes"])
        return result

    return jsonify({
        "project_name": state.project.meta.name,
        "total_scenes": total_scenes,
        "episodes": episodes_info,
        "characters": _fmt(chars),
        "locations": _fmt(locations),
        "props": _fmt(props),
    })


# ── Project Archive Export ────────────────────────────────────────

@app.route("/api/projects/current/archive")
def export_project_archive():
    """Pack all project data into a .zip archive for collaboration."""
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    # Persist current episode before archiving
    state.save_to_episode()
    state.project._save_meta()

    project_dir = state.project.project_dir

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Project metadata
        meta_path = os.path.join(project_dir, "project.json")
        if os.path.exists(meta_path):
            zf.write(meta_path, "project.json")

        # All episode files
        for fname in os.listdir(project_dir):
            if fname.startswith("episode_") and fname.endswith(".json"):
                zf.write(os.path.join(project_dir, fname), fname)

    buf.seek(0)
    safe_name = state.project.meta.name.replace("/", "_").replace("\\", "_")
    filename = f"{safe_name}_project.sbp"   # .sbp = Script Breakdown Project
    return Response(
        buf.getvalue(),
        mimetype="application/zip",
        headers={"Content-Disposition": _make_content_disposition(filename)},
    )


# ── Project Archive Import ────────────────────────────────────────

@app.route("/api/projects/import", methods=["POST"])
def import_project_archive():
    """Restore a project from a .sbp archive file."""
    if "file" not in request.files:
        return jsonify({"error": "没有上传文件"}), 400

    file = request.files["file"]
    if not (file.filename.endswith(".sbp") or file.filename.endswith(".zip")):
        return jsonify({"error": "请上传 .sbp 或 .zip 格式的工程文件"}), 400

    try:
        raw = io.BytesIO(file.read())
        with zipfile.ZipFile(raw, "r") as zf:
            names = zf.namelist()
            if "project.json" not in names:
                return jsonify({"error": "文件格式不正确：缺少 project.json"}), 400

            # Parse original meta to get project name
            original_meta = json.loads(zf.read("project.json").decode("utf-8"))
            import uuid as _uuid
            new_id = str(_uuid.uuid4())[:8]
            new_dir = os.path.join(PROJECTS_DIR, new_id)
            os.makedirs(new_dir, exist_ok=True)

            # Rewrite project.json with new ID, appending "(导入)" to name
            original_meta["id"] = new_id
            original_meta["name"] = original_meta.get("name", "未命名项目") + " (导入)"
            with open(os.path.join(new_dir, "project.json"), "w", encoding="utf-8") as f:
                json.dump(original_meta, f, ensure_ascii=False, indent=2)

            # Extract episode files
            for name in names:
                if name.startswith("episode_") and name.endswith(".json"):
                    data = zf.read(name)
                    with open(os.path.join(new_dir, name), "wb") as f:
                        f.write(data)

        # Open the newly imported project
        from src.web.project import Project as _Project
        state = AppState()
        state.reset()
        project = _Project.load(new_id)
        state.project = project

        if project.meta.active_episode_id:
            try:
                state.load_from_episode(project.meta.active_episode_id)
            except Exception:
                pass

        state.status_message = f"工程已导入: {project.meta.name}"
        _save_last_project(project.meta.id)
        return jsonify(state.to_dict())

    except zipfile.BadZipFile:
        return jsonify({"error": "文件损坏或格式不正确"}), 400
    except Exception as e:
        return jsonify({"error": f"导入失败: {e}"}), 500


# ── Schedule Management ──────────────────────────────────────────

def _collect_all_scenes(state: AppState) -> tuple:
    """Collect scenes + entity data across all episodes (or current state if no project).

    Returns (all_scenes: list[Scene], entities_by_scene_number: dict).
    When a project is open the current episode is flushed first, then every
    episode is iterated in order and scenes are re-numbered globally (1, 2, 3…)
    so that scene_number is unique across the full collection.
    """
    if state.project:
        state.save_to_episode()
        all_scenes: list = []
        entities_by_num: dict = {}
        global_num = 1
        for ep_id in state.project.meta.episode_order:
            try:
                ep = state.project.load_episode(ep_id)
            except Exception:
                continue
            ep_entities = ep.entities or {}
            for i, sd in enumerate(ep.scenes or []):
                scene = Scene(
                    scene_number=global_num,
                    heading=sd.get("heading", ""),
                    location=sd.get("location", ""),
                    int_ext=sd.get("int_ext", ""),
                    time_of_day=sd.get("time_of_day", ""),
                    start_line=sd.get("start_line", 0),
                    end_line=sd.get("end_line", 0),
                )
                all_scenes.append(scene)
                ent = ep_entities.get(i) or ep_entities.get(str(i)) or {}
                entities_by_num[global_num] = ent
                global_num += 1
        return all_scenes, entities_by_num

    # No project: use the currently loaded state
    if state.scene_list:
        return list(state.scene_list), _entities_by_scene_number(state)
    return [], {}


def _get_duration_params(state) -> DurationParams:
    """加载项目的时长参数（用户可自定义）"""
    if not state.project:
        return DurationParams()
    path = os.path.join(state.project.project_dir, "duration_params.json")
    return DurationParams.load(path)


def _save_duration_params(state, params: DurationParams) -> None:
    if state.project:
        path = os.path.join(state.project.project_dir, "duration_params.json")
        params.save(path)


def _ensure_schedule_loaded(state: AppState) -> None:
    """Auto-load schedule + snapshots from project if not yet in memory."""
    if state.schedule is None and state.project:
        loaded = state.project.load_schedule()
        if loaded is not None:
            state.schedule = loaded
    if not state.schedule_snapshots and state.project:
        snaps_meta = state.project.list_schedule_snapshots()
        for meta in snaps_meta:
            try:
                snap = state.project.load_schedule_snapshot(meta["version"])
                state.schedule_snapshots.append(snap)
            except Exception:
                pass


@app.route("/api/schedule", methods=["GET"])
def get_schedule():
    """Return the current ProductionSchedule (auto-loaded from project if needed)."""
    state = AppState()
    _ensure_schedule_loaded(state)
    if state.schedule is None:
        return jsonify(None)
    return jsonify(state.schedule.to_dict())


@app.route("/api/schedule", methods=["PUT"])
def put_schedule():
    """Replace the current schedule with the request body and persist."""
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400
    data = request.get_json()
    if not data:
        return jsonify({"error": "请求体不能为空"}), 400
    state.schedule = ProductionSchedule.from_dict(data)
    _persist_schedule(state)
    return jsonify(state.schedule.to_dict())


@app.route("/api/schedule/duration-params", methods=["GET", "PUT"])
def duration_params_api():
    """获取或更新时长估算参数（表A/B/C + 项目类型）"""
    state = AppState()
    if request.method == "GET":
        p = _get_duration_params(state)
        return jsonify(p.to_dict())
    data = request.get_json() or {}
    p = DurationParams.from_dict(data)
    _save_duration_params(state, p)
    return jsonify({"ok": True})


@app.route("/api/schedule/estimate-durations")
def estimate_durations():
    """预览所有场次的时长估算（不改变状态）"""
    state = AppState()
    all_scenes, entities_by_num = _collect_all_scenes(state)
    if not all_scenes:
        return jsonify({"error": "未找到场次"}), 400
    params = _get_duration_params(state)
    # 允许前端传 genre_key 覆盖
    genre_key = request.args.get("genre_key")
    if genre_key:
        params.genre_key = genre_key
    engine = SceneDurationEngine(params)
    estimates = engine.estimate_all(all_scenes, entities_by_num)
    vals = [e["total_hours"] for e in estimates.values()]
    total_h = sum(vals)
    return jsonify({
        "scene_count": len(vals),
        "total_hours": round(total_h, 1),
        "avg_hours": round(total_h / max(len(vals), 1), 2),
        "min_days_10h": max(1, math.ceil(total_h / 10)),
        "min_days_12h": max(1, math.ceil(total_h / 12)),
        "genre_key": params.genre_key,
        "genre_factor": params.genre_factor,
        "estimates": {str(k): v for k, v in estimates.items()},
    })


@app.route("/api/scenes/<int:scene_num>/duration")
def scene_duration_detail(scene_num: int):
    """返回单场的时长估算详情，包括公式分解"""
    state = AppState()
    all_scenes, entities_by_num = _collect_all_scenes(state)
    scene = next((s for s in all_scenes if s.scene_number == scene_num), None)
    if not scene:
        return jsonify({"error": "场次不存在"}), 404
    params = _get_duration_params(state)
    engine = SceneDurationEngine(params)
    ent = entities_by_num.get(scene_num, {})
    detail = engine.estimate_scene(scene, ent)
    detail["custom_factors"] = ent.get("custom_factors", [])
    return jsonify(detail)


# ── Mod 5b: 辅助函数 ──────────────────────────────────────────────────────────

def _save_entities_to_episodes(state: AppState, updates: dict) -> None:
    """将 {global_scene_number: {field: value}} 的更新写回各集 entities 并保存。

    有 project 时写到各集文件；无 project 时将更新直接写到 state.entities（按本地索引）。
    """
    if not state.project:
        # 非项目模式：scene_number → local index
        if state.scene_list:
            for i, sc in enumerate(state.scene_list):
                if sc.scene_number in updates:
                    ent = dict(state.entities.get(i) or {})
                    ent.update(updates[sc.scene_number])
                    state.entities[i] = ent
        return
    state.save_to_episode()  # 确保当前集已落盘

    global_num = 1
    for ep_id in state.project.meta.episode_order:
        try:
            ep = state.project.load_episode(ep_id)
        except Exception:
            continue
        changed = False
        for i in range(len(ep.scenes or [])):
            if global_num in updates:
                ent = ep.entities.get(i) or {}
                ent.update(updates[global_num])
                ep.entities[i] = ent
                changed = True
            global_num += 1
        if changed:
            state.project.save_episode(ep)


# ── Mod 5a: LLM 批量场景类型分类 ─────────────────────────────────────────────

_llm_classify_status: dict = {}   # task_id → {status, progress, total, results, error}


@app.route("/api/scenes/llm/classify-types", methods=["POST"])
def llm_classify_scene_types():
    """POST /api/scenes/llm/classify-types
    异步使用 LLM 批量识别所有场次的类型，每批 30 场。
    返回 task_id，前端轮询 /api/scenes/llm/classify-status/<task_id>。
    可选 JSON 体：{"genre_key": "B5_古装"}
    """
    import uuid
    state = AppState()
    data = request.get_json() or {}
    genre_key = data.get("genre_key")

    all_scenes, entities_by_num = _collect_all_scenes(state)
    if not all_scenes:
        return jsonify({"error": "未找到场次"}), 400

    try:
        llm = _create_llm(state.llm_config)
    except Exception as e:
        return jsonify({"error": f"LLM 未配置: {e}"}), 400

    task_id = str(uuid.uuid4())[:8]
    _llm_classify_status[task_id] = {
        "status": "running", "progress": 0,
        "total": len(all_scenes), "results": {}, "error": None,
    }

    params = _get_duration_params(state)
    if genre_key:
        params.genre_key = genre_key
    table_a_keys = list(params.table_a.keys()) + list(params.custom_scene_types.keys())

    def _worker():
        try:
            batch_size = 30
            all_results: dict = {}
            scenes_list = all_scenes  # captured from outer scope
            for batch_start in range(0, len(scenes_list), batch_size):
                batch = scenes_list[batch_start: batch_start + batch_size]
                # 构建 prompt
                lines = []
                for s in batch:
                    ent = entities_by_num.get(s.scene_number, {})
                    chars = ent.get("characters", [])
                    lines.append(
                        f"场次{s.scene_number}: 标题={s.heading!r}, "
                        f"角色={chars}, 摘要={ent.get('summary','')[:80]}"
                    )
                scene_desc = "\n".join(lines)
                type_list = "\n".join(f"- {k}" for k in table_a_keys)
                prompt = (
                    f"你是影视制片助手，请将以下场次分类到场景类型中。\n"
                    f"可选类型（只能从中选一个）：\n{type_list}\n\n"
                    f"场次信息：\n{scene_desc}\n\n"
                    f"请以 JSON 格式返回，格式为：\n"
                    f'{{"场次号": "类型键", ...}}\n'
                    f"例如：{{\"1\": \"A7_武戏\", \"2\": \"A1_简单对话\"}}\n"
                    f"只返回 JSON，不要其他说明。"
                )
                try:
                    raw = llm.complete(prompt)
                    # 提取 JSON 部分
                    import re
                    m = re.search(r'\{[^{}]*\}', raw, re.DOTALL)
                    if m:
                        batch_result = json.loads(m.group())
                        for sn_str, type_key in batch_result.items():
                            try:
                                sn = int(sn_str)
                            except ValueError:
                                continue
                            if type_key in table_a_keys:
                                all_results[sn] = {"scene_type_key": type_key}
                except Exception:
                    pass  # 单批失败不影响其他批次
                _llm_classify_status[task_id]["progress"] = batch_start + len(batch)

            # 写回 entities
            _save_entities_to_episodes(state, all_results)
            _llm_classify_status[task_id].update({
                "status": "done", "results": all_results,
                "progress": len(scenes_list),
            })
        except Exception as exc:
            _llm_classify_status[task_id].update(
                {"status": "error", "error": str(exc)}
            )

    threading.Thread(target=_worker, daemon=True).start()
    return jsonify({"task_id": task_id, "total": len(all_scenes)})


@app.route("/api/scenes/llm/classify-status/<task_id>")
def llm_classify_status(task_id: str):
    """轮询 LLM 批量分类状态"""
    info = _llm_classify_status.get(task_id)
    if not info:
        return jsonify({"error": "任务不存在"}), 404
    return jsonify(info)


# ── Mod 6: 自定义场景类型 CRUD ────────────────────────────────────────────────

@app.route("/api/schedule/custom-scene-types", methods=["GET", "POST", "DELETE"])
def custom_scene_types_api():
    """
    GET  → 返回当前所有自定义场景类型
    POST → 新增/更新一个自定义类型（body: {key, label, minutes, keywords, priority}）
    DELETE → 删除一个自定义类型（body: {key}）
    """
    state = AppState()
    params = _get_duration_params(state)

    if request.method == "GET":
        return jsonify(params.custom_scene_types)

    data = request.get_json() or {}
    key = str(data.get("key", "")).strip()
    if not key:
        return jsonify({"error": "缺少 key"}), 400

    if request.method == "DELETE":
        params.custom_scene_types.pop(key, None)
        # 同时从 keyword_rules 移除（若存在）
        params.keyword_rules.pop(key, None)
        _save_duration_params(state, params)
        return jsonify({"ok": True})

    # POST：新增/更新
    label = str(data.get("label", key))
    minutes = int(data.get("minutes", 20))
    keywords = data.get("keywords", [])
    priority = int(data.get("priority", 50))
    params.custom_scene_types[key] = {"label": label, "minutes": minutes}
    if keywords:
        params.keyword_rules[key] = {
            "keywords": keywords,
            "priority": priority,
            "description": label,
        }
    _save_duration_params(state, params)
    return jsonify({"ok": True, "key": key})


@app.route("/api/scenes/type-override", methods=["POST"])
def scene_type_override():
    """POST /api/scenes/type-override
    持久化单个场次的手动类型指定。
    body: {scene_number: int, scene_type_key: str}
    """
    state = AppState()
    data = request.get_json() or {}
    scene_num = data.get("scene_number")
    type_key = str(data.get("scene_type_key", "") or "").strip()
    if scene_num is None:
        return jsonify({"error": "缺少 scene_number"}), 400
    _save_entities_to_episodes(state, {int(scene_num): {"scene_type_key": type_key}})
    return jsonify({"ok": True})


# ── 主页面：场次时长 & 类型分类 ───────────────────────────────────────────────

@app.route("/api/scenes/<int:idx>/duration-by-index")
def scene_duration_by_index(idx: int):
    """返回本地索引 idx 场次的时长估算详情（主页面用）。"""
    state = AppState()
    if not state.scene_list or idx < 0 or idx >= len(state.scene_list):
        return jsonify({"error": "无效场次索引"}), 400
    scene = state.scene_list[idx]
    params = _get_duration_params(state)
    engine = SceneDurationEngine(params)
    ent = (state.entities or {}).get(idx) or {}
    # 如果有用户手动覆盖时长，也一并返回
    detail = engine.estimate_scene(scene, ent)
    detail["duration_override_hours"] = ent.get("duration_override_hours")
    detail["custom_factors"] = ent.get("custom_factors", [])
    return jsonify(detail)


@app.route("/api/scenes/<int:idx>/duration-override", methods=["PUT"])
def scene_duration_override(idx: int):
    """保存单个场次的时长覆盖和自定义因子。

    Body:
      duration_override_hours: float | null
      custom_factors: [{"name": str, "value": float}, ...]
      scene_type_key: str | null
    """
    state = AppState()
    if not state.scene_list or idx < 0 or idx >= len(state.scene_list):
        return jsonify({"error": "无效场次索引"}), 400

    data = request.get_json() or {}
    ent = dict((state.entities or {}).get(idx) or {})

    if "duration_override_hours" in data:
        val = data["duration_override_hours"]
        ent["duration_override_hours"] = float(val) if val not in (None, "", "null") else None

    if "custom_factors" in data:
        factors = data["custom_factors"]
        if isinstance(factors, list):
            ent["custom_factors"] = [
                {"name": str(f["name"]), "value": float(f["value"])}
                for f in factors
                if isinstance(f, dict) and "name" in f and "value" in f
            ]

    if "scene_type_key" in data:
        val = data["scene_type_key"]
        ent["scene_type_key"] = str(val) if val else ""

    state.entities[idx] = ent
    state.status_message = f"场次 {idx + 1} 时长参数已更新"
    return jsonify(state.to_dict())


@app.route("/api/scenes/<int:idx>/classify-type", methods=["POST"])
def classify_scene_type(idx: int):
    """规则快速分类单个场次类型，结果写入 entities[idx]，返回 state。"""
    state = AppState()
    if not state.scene_list or idx < 0 or idx >= len(state.scene_list):
        return jsonify({"error": "无效场次索引"}), 400
    scene = state.scene_list[idx]
    params = _get_duration_params(state)
    engine = SceneDurationEngine(params)
    ent = dict((state.entities or {}).get(idx) or {})
    cls = engine.classify_scene(scene, ent)
    ent["scene_type_key"] = cls["a_key"]
    ent["classification"] = cls
    state.entities[idx] = ent
    return jsonify(state.to_dict())


@app.route("/api/scenes/classify-all", methods=["POST"])
def classify_all_scene_types():
    """规则批量分类所有场次类型，写入 entities，返回 state。"""
    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "未找到场次"}), 400
    params = _get_duration_params(state)
    engine = SceneDurationEngine(params)
    for i, scene in enumerate(state.scene_list):
        ent = dict((state.entities or {}).get(i) or {})
        cls = engine.classify_scene(scene, ent)
        ent["scene_type_key"] = cls["a_key"]
        ent["classification"] = cls
        state.entities[i] = ent
    return jsonify(state.to_dict())


@app.route("/api/schedule/generate", methods=["POST"])
def schedule_generate():
    """Generate an initial schedule synchronously using the greedy heuristic.

    Traverses ALL episodes in the project (or the current scene list if no
    project is open).  The result is the greedy initial solution only — call
    POST /api/schedule/optimize afterwards for the SA local-search pass.

    Body (all optional except start_date):
      start_date, max_hours_per_day, rest_days,
      weight_transition, weight_actor, weight_location, weight_balance, weight_days
    """
    state = AppState()
    if not state.project and not state.scene_list:
        return jsonify({"error": "请先打开项目或导入剧本"}), 400

    data = request.get_json() or {}
    start_date = data.get("start_date", "")
    if not start_date:
        return jsonify({"error": "缺少 start_date"}), 400

    # Use learned weights as defaults if available and the user didn't override them
    learner = _get_learner(state)
    base_config = ScheduleConfig(
        start_date=start_date,
        max_hours_per_day=float(data.get("max_hours_per_day", 12.0)),
        rest_days=data.get("rest_days", []),
        weight_transition=1.0,
        weight_actor=1.0,
        weight_location=1.0,
        weight_balance=0.5,
        weight_days=1.5,
    )
    if learner and not any(
        k in data for k in (
            "weight_transition", "weight_actor", "weight_location",
            "weight_balance", "weight_days",
        )
    ):
        base_config = learner.get_suggested_config(base_config)

    config = ScheduleConfig(
        start_date=start_date,
        max_hours_per_day=float(data.get("max_hours_per_day", base_config.max_hours_per_day)),
        rest_days=data.get("rest_days", list(base_config.rest_days)),
        weight_transition=float(data.get("weight_transition", base_config.weight_transition)),
        weight_actor=float(data.get("weight_actor", base_config.weight_actor)),
        weight_location=float(data.get("weight_location", base_config.weight_location)),
        weight_balance=float(data.get("weight_balance", base_config.weight_balance)),
        weight_days=float(data.get("weight_days", base_config.weight_days)),
    )

    all_scenes, entities_by_num = _collect_all_scenes(state)
    if not all_scenes:
        return jsonify({"error": "未找到任何场次数据，请先完成场次识别"}), 400

    # ── 估算场次时长（使用新引擎） ──
    duration_params = _get_duration_params(state)
    duration_params.genre_key = data.get("genre_key", duration_params.genre_key)
    engine = SceneDurationEngine(duration_params)
    config.scene_duration_map = engine.get_duration_map(all_scenes, entities_by_num)
    config.genre_key = duration_params.genre_key

    # ── 根据场次规模自动调整求解时间 ──
    n = len(all_scenes)
    config.solver_time_limit = int(data.get(
        "solver_time_limit",
        60 if n > 150 else 30 if n > 50 else 15,
    ))
    config.constraint_level = data.get("constraint_level", "relaxed")

    actors = state.schedule.actors if state.schedule else []
    locations = state.schedule.locations if state.schedule else []

    optimizer = ScheduleOptimizer(
        scenes=all_scenes,
        entities=entities_by_num,
        actors=actors,
        locations=locations,
        config=config,
    )
    # generate_safe never raises — falls back through relaxation → greedy → sequential
    result = optimizer.generate_safe(time_limit_seconds=10)

    if state.schedule is not None:
        _save_schedule_snapshot(state, "before_generate")
        state.schedule.config = config
    else:
        state.schedule = ProductionSchedule(
            actors=actors,
            locations=locations,
            config=config,
        )
    state.schedule.shooting_days = result.shooting_days
    state.schedule.start_date = result.start_date
    state.schedule.end_date = result.end_date
    _persist_schedule(state)
    resp = state.schedule.to_dict()
    # Surface warnings and solver status to the frontend
    resp["warnings"] = result.warnings
    resp["solver_status"] = result.solver_status
    relaxation_note = getattr(result, "_relaxation_note", None)
    if relaxation_note:
        resp["warning"] = relaxation_note
    return jsonify(resp)


@app.route("/api/schedule/optimize", methods=["POST"])
def schedule_optimize():
    """Run simulated-annealing local-search on the existing schedule asynchronously.

    Uses state.llm_task to track progress.  Poll GET /api/llm/status (or
    GET /api/schedule) to detect completion.
    """
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule or not state.schedule.shooting_days:
        return jsonify({"error": "暂无排期数据，请先调用 generate"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有任务运行中"}), 409

    all_scenes, entities_by_num = _collect_all_scenes(state)
    if not all_scenes:
        return jsonify({"error": "未找到场次数据"}), 400

    state.llm_task.task_type = "optimize_schedule"
    state.llm_task.status = "running"
    state.llm_task.error = ""
    state.llm_task.result = None

    current_schedule = state.schedule

    def _run():
        try:
            optimizer = ScheduleOptimizer(
                scenes=all_scenes,
                entities=entities_by_num,
                actors=current_schedule.actors,
                locations=current_schedule.locations,
                config=current_schedule.config,
            )
            learner = _get_learner(state)
            extra = learner.get_extra_constraints() if learner else None
            # generate_safe never raises — degrades gracefully through fallbacks
            improved = optimizer.generate_safe(time_limit_seconds=30, extra_constraints=extra or None)
            _save_schedule_snapshot(state, "before_optimize")
            state.schedule.shooting_days = improved.shooting_days
            state.schedule.end_date = improved.end_date
            _persist_schedule(state)
            task_result: dict = {
                "warnings": improved.warnings,
                "solver_status": improved.solver_status,
            }
            relaxation_note = getattr(improved, "_relaxation_note", None)
            if relaxation_note:
                task_result["warning"] = relaxation_note
            state.llm_task.result = task_result
            state.llm_task.status = "done"
        except Exception as e:
            state.llm_task.status = "error"
            state.llm_task.error = str(e)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "status": "running"})


@app.route("/api/schedule/llm-loop", methods=["POST"])
def schedule_llm_loop():
    """Run the CP-SAT + LLM closed-loop refinement asynchronously.

    The solver produces an initial schedule, the LLM reviews it and suggests
    constraint improvements, and the solver re-runs — up to *max_rounds* times.

    Body (all optional):
      time_limit : int   — seconds per CP-SAT round (default 15)
      max_rounds : int   — maximum feedback iterations (default 3)

    Progress is tracked via state.llm_task; poll GET /api/llm/status.
    When done, state.schedule is updated and the round history is stored in
    state.llm_task.result as a list of per-round dicts.
    """
    state = AppState()
    _ensure_schedule_loaded(state)
    if state.llm_task.status == "running":
        return jsonify({"error": "已有任务运行中"}), 409

    all_scenes, entities_by_num = _collect_all_scenes(state)
    if not all_scenes:
        return jsonify({"error": "未找到场次数据，请先完成场次识别"}), 400

    data        = request.get_json() or {}
    time_limit  = int(data.get("time_limit",  15))
    max_rounds  = int(data.get("max_rounds",   3))

    # Determine schedule config — use existing schedule config if available,
    # otherwise build a minimal default.
    if state.schedule and state.schedule.config.start_date:
        sched_config = state.schedule.config
        actors    = list(state.schedule.actors)
        locations = list(state.schedule.locations)
    else:
        return jsonify({"error": "暂无排期数据，请先调用 generate 或 optimize"}), 400

    state.llm_task.task_type = "llm_loop"
    state.llm_task.status    = "running"
    state.llm_task.error     = ""
    state.llm_task.result    = None

    def _run() -> None:
        try:
            optimizer = ScheduleOptimizer(
                scenes=all_scenes,
                entities=entities_by_num,
                actors=actors,
                locations=locations,
                config=sched_config,
            )
            llm = _create_llm(state.llm_config)
            loop = LLMScheduleLoop(optimizer, llm, max_rounds=max_rounds)

            learner = _get_learner(state)
            initial = learner.get_extra_constraints() if learner else None
            result = loop.run(
                time_limit_per_round=time_limit,
                initial_constraints=initial or None,
            )

            new_schedule: ProductionSchedule = result["schedule"]
            _save_schedule_snapshot(state, "before_llm_loop")
            state.schedule.shooting_days = new_schedule.shooting_days
            state.schedule.end_date      = new_schedule.end_date
            _persist_schedule(state)

            # Persist LLM-suggested constraints into the preference learner
            if learner:
                for round_rec in result.get("rounds", []):
                    for issue in round_rec.get("issues_raw", []):
                        c = issue.get("constraint")
                        if c and isinstance(c, dict):
                            learner.record_rule_from_llm({
                                "type":   c.get("type", ""),
                                "scenes": c.get("scenes", []),
                                "params": c.get("params") or {},
                                "reason": issue.get("description", ""),
                            })

            # Store the per-round history so the frontend can display reasoning
            relaxation_note = getattr(new_schedule, "_relaxation_note", None)
            state.llm_task.result = {
                "rounds":       result["rounds"],
                "total_rounds": result["total_rounds"],
                "final_score":  result["final_score"],
                "warning":      relaxation_note,
            }
            state.llm_task.status = "done"
        except Exception as exc:
            state.llm_task.status = "error"
            state.llm_task.error  = str(exc)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "status": "running", "max_rounds": max_rounds})


@app.route("/api/schedule/conflicts", methods=["GET"])
def schedule_conflicts():
    """Return all hard-constraint violations for the current schedule (synchronous)."""
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify([])

    all_scenes, entities_by_num = _collect_all_scenes(state)
    checker = ConstraintChecker(
        all_scenes,
        entities_by_num,
        state.schedule.actors,
        state.schedule.locations,
    )
    violations = checker.check_all(state.schedule)
    return jsonify([
        {
            "type": v.type,
            "severity": v.severity,
            "message": v.message,
            "day_date": v.day_date,
            "scene_ids": v.scene_ids,
        }
        for v in violations
    ])


@app.route("/api/schedule/reschedule", methods=["POST"])
def schedule_reschedule():
    """Incremental reschedule triggered by a change event.

    Saves a snapshot before running, then optimises asynchronously.
    Body: { "change_type": str, "change_data": dict }
    change_type: "weather" | "actor" | "script_add" | "script_remove" | "location"
    """
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有任务运行中"}), 409

    data = request.get_json() or {}
    change_type = data.get("change_type", "")
    change_data = data.get("change_data", {})
    if not change_type:
        return jsonify({"error": "缺少 change_type"}), 400

    all_scenes, entities_by_num = _collect_all_scenes(state)
    if not all_scenes:
        return jsonify({"error": "未找到场次数据"}), 400

    # Save snapshot before mutating
    _save_schedule_snapshot(state, f"before_reschedule_{change_type}")

    state.llm_task.task_type = "reschedule"
    state.llm_task.status = "running"
    state.llm_task.error = ""
    state.llm_task.result = None

    current_schedule = state.schedule

    def _run():
        try:
            optimizer = ScheduleOptimizer(
                scenes=all_scenes,
                entities=entities_by_num,
                actors=current_schedule.actors,
                locations=current_schedule.locations,
                config=current_schedule.config,
            )
            result = optimizer.reschedule(current_schedule, change_type, change_data)
            new_sched = result["schedule"]
            state.schedule.shooting_days = new_sched.shooting_days
            state.schedule.end_date = new_sched.end_date
            _persist_schedule(state)
            state.llm_task.status = "done"
            state.llm_task.result = result["impact"]
        except Exception as e:
            state.llm_task.status = "error"
            state.llm_task.error = str(e)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "status": "running"})


@app.route("/api/schedule/days/<date>", methods=["PUT"])
def schedule_days_update(date):
    """Update metadata for a shooting day (notes, call_time, location, etc.)."""
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 404

    day = next((d for d in state.schedule.shooting_days if d.date == date), None)
    if day is None:
        return jsonify({"error": f"找不到拍摄日: {date}"}), 404

    data = request.get_json() or {}
    # Capture snapshot before mutation for preference learning (location changes)
    if "location" in data:
        import copy as _copy
        before_schedule = _copy.deepcopy(state.schedule)
    else:
        before_schedule = None

    for field in ("location", "call_time", "estimated_end", "notes", "weather_backup"):
        if field in data:
            setattr(day, field, data[field])

    _persist_schedule(state)
    if before_schedule is not None:
        _record_schedule_adjustment(state, before_schedule, state.schedule)
    return jsonify(day.to_dict())


@app.route("/api/schedule/days/<date>/add_scene", methods=["POST"])
def schedule_days_add_scene(date):
    """Add a scene to a shooting day.  Body: { "scene_id": int }"""
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 404

    day = next((d for d in state.schedule.shooting_days if d.date == date), None)
    if day is None:
        return jsonify({"error": f"找不到拍摄日: {date}"}), 404

    data = request.get_json() or {}
    scene_id = data.get("scene_id")
    if scene_id is None:
        return jsonify({"error": "缺少 scene_id"}), 400

    import copy as _copy
    before_schedule = _copy.deepcopy(state.schedule)
    scene_id = int(scene_id)
    if scene_id not in day.scene_ids:
        day.scene_ids.append(scene_id)
    _persist_schedule(state)
    _record_schedule_adjustment(state, before_schedule, state.schedule)
    return jsonify(day.to_dict())


@app.route("/api/schedule/days/<date>/remove_scene", methods=["POST"])
def schedule_days_remove_scene(date):
    """Remove a scene from a shooting day.  Body: { "scene_id": int }"""
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 404

    day = next((d for d in state.schedule.shooting_days if d.date == date), None)
    if day is None:
        return jsonify({"error": f"找不到拍摄日: {date}"}), 404

    data = request.get_json() or {}
    scene_id = data.get("scene_id")
    if scene_id is None:
        return jsonify({"error": "缺少 scene_id"}), 400

    import copy as _copy
    before_schedule = _copy.deepcopy(state.schedule)
    scene_id = int(scene_id)
    day.scene_ids = [s for s in day.scene_ids if s != scene_id]
    _persist_schedule(state)
    _record_schedule_adjustment(state, before_schedule, state.schedule)
    return jsonify(day.to_dict())


@app.route("/api/schedule/days/<date>/status", methods=["PUT"])
def schedule_days_status(date):
    """Update the shooting status of a day.  Body: { "status": str }"""
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 404

    day = next((d for d in state.schedule.shooting_days if d.date == date), None)
    if day is None:
        return jsonify({"error": f"找不到拍摄日: {date}"}), 404

    data = request.get_json() or {}
    status = data.get("status", "")
    if status not in ("planned", "shooting", "completed", "cancelled"):
        return jsonify({"error": "无效状态，允许值: planned/shooting/completed/cancelled"}), 400

    day.status = status
    _persist_schedule(state)
    return jsonify(day.to_dict())


@app.route("/api/schedule/actors", methods=["GET"])
def get_schedule_actors():
    """Return actors merged from entity data and existing ActorSchedule objects.

    Characters found across all episodes are included even if no ActorSchedule
    exists yet.  The extra field 'scene_count' shows total scene appearances.
    """
    state = AppState()
    _ensure_schedule_loaded(state)

    _, entities_by_num = _collect_all_scenes(state)

    # Count scenes per character
    char_scene_counts: dict = {}
    for ent in entities_by_num.values():
        for ch in ent.get("characters", []):
            char_scene_counts[ch] = char_scene_counts.get(ch, 0) + 1

    # Index existing ActorSchedule objects
    actor_map: dict = {a.character_name: a for a in (state.schedule.actors if state.schedule else [])}

    result = []
    seen: set = set()
    for char, count in sorted(char_scene_counts.items(), key=lambda x: -x[1]):
        seen.add(char)
        entry = (actor_map[char].to_dict() if char in actor_map
                 else ActorSchedule(character_name=char).to_dict())
        entry["scene_count"] = count
        result.append(entry)

    # Include ActorSchedule entries not found in entity data
    for char, actor in actor_map.items():
        if char not in seen:
            entry = actor.to_dict()
            entry["scene_count"] = 0
            result.append(entry)

    return jsonify(result)


@app.route("/api/schedule/actors/<path:name>", methods=["PUT"])
def put_schedule_actor(name):
    """Create or update an ActorSchedule by character name."""
    state = AppState()
    if state.schedule is None:
        state.schedule = ProductionSchedule(config=ScheduleConfig(start_date=""))

    data = request.get_json() or {}
    data["character_name"] = name
    actor = ActorSchedule.from_dict(data)
    state.schedule.actors = [a for a in state.schedule.actors if a.character_name != name]
    state.schedule.actors.append(actor)
    _persist_schedule(state)
    return jsonify(actor.to_dict())


@app.route("/api/schedule/locations", methods=["GET"])
def get_schedule_locations():
    """Return locations merged from scene data and existing LocationInfo objects.

    Locations found across all episodes are included even if no LocationInfo
    exists yet.  The extra field 'scene_count' shows total scene count.
    """
    state = AppState()
    _ensure_schedule_loaded(state)

    all_scenes, _ = _collect_all_scenes(state)

    # Count scenes per location
    loc_scene_counts: dict = {}
    for scene in all_scenes:
        loc = (scene.location or "").strip()
        if loc:
            loc_scene_counts[loc] = loc_scene_counts.get(loc, 0) + 1

    # Index existing LocationInfo objects
    loc_map: dict = {l.name: l for l in (state.schedule.locations if state.schedule else [])}

    result = []
    seen: set = set()
    for loc, count in sorted(loc_scene_counts.items(), key=lambda x: -x[1]):
        seen.add(loc)
        entry = (loc_map[loc].to_dict() if loc in loc_map
                 else LocationInfo(name=loc).to_dict())
        entry["scene_count"] = count
        result.append(entry)

    for loc_name, loc_info in loc_map.items():
        if loc_name not in seen:
            entry = loc_info.to_dict()
            entry["scene_count"] = 0
            result.append(entry)

    return jsonify(result)


@app.route("/api/schedule/locations/<path:name>", methods=["PUT"])
def put_schedule_location(name):
    """Create or update a LocationInfo by location name."""
    state = AppState()
    if state.schedule is None:
        state.schedule = ProductionSchedule(config=ScheduleConfig(start_date=""))

    data = request.get_json() or {}
    data["name"] = name
    loc = LocationInfo.from_dict(data)
    state.schedule.locations = [l for l in state.schedule.locations if l.name != name]
    state.schedule.locations.append(loc)
    _persist_schedule(state)
    return jsonify(loc.to_dict())


@app.route("/api/schedule/snapshots", methods=["GET"])
def schedule_snapshots():
    """List all schedule snapshots (loads from project disk if memory is empty)."""
    state = AppState()
    _ensure_schedule_loaded(state)
    return jsonify([
        {
            "version": s.version,
            "timestamp": s.timestamp,
            "trigger": s.trigger,
            "diff_summary": s.diff_summary,
        }
        for s in state.schedule_snapshots
    ])


@app.route("/api/schedule/snapshots/<int:version>/restore", methods=["POST"])
def restore_schedule_snapshot(version):
    """Roll back the schedule to a specific snapshot version."""
    state = AppState()
    _ensure_schedule_loaded(state)

    snap = next((s for s in state.schedule_snapshots if s.version == version), None)
    if snap is None and state.project:
        try:
            snap = state.project.load_schedule_snapshot(version)
        except FileNotFoundError:
            pass
    if snap is None:
        return jsonify({"error": f"快照版本 {version} 不存在"}), 404

    _save_schedule_snapshot(state, "before_rollback")
    state.schedule = ProductionSchedule.from_dict(snap.schedule_data)
    _persist_schedule(state)
    return jsonify(state.schedule.to_dict())


# ── Export ───────────────────────────────────────────────────────

@app.route("/api/export/<fmt>")
def export(fmt):
    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "无场次数据可导出"}), 400

    if fmt == "txt":
        return _export_txt(state)
    elif fmt == "csv":
        return _export_csv(state)
    else:
        return jsonify({"error": f"不支持的导出格式: {fmt}"}), 400


def _export_txt(state: AppState) -> Response:
    lines = []
    for scene in state.scene_list:
        lines.append("=" * 60)
        lines.append(f"场次 {scene.scene_number}: {scene.heading}")
        lines.append(f"内外景: {scene.int_ext}  地点: {scene.location}  时间: {scene.time_of_day}")
        lines.append(f"行范围: {scene.start_line + 1}-{scene.end_line}")
        if scene.summary:
            lines.append(f"摘要: {scene.summary}")

        # Include entity info if available
        idx = scene.scene_number - 1
        if idx in state.entities:
            ent = state.entities[idx]
            if ent.get("characters"):
                lines.append(f"人物: {', '.join(ent['characters'])}")
            if ent.get("props"):
                lines.append(f"道具: {', '.join(ent['props'])}")

        lines.append("=" * 60)
        lines.append("")

    content = "\n".join(lines)
    filename = os.path.splitext(state.filename or "export")[0] + "_scenes.txt"
    return Response(
        content.encode("utf-8"),
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": _make_content_disposition(filename)},
    )


def _export_csv(state: AppState) -> Response:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["场次号", "标题", "内外景", "地点", "时间", "起始行", "结束行",
                      "置信度", "人工校准", "摘要", "人物", "道具"])
    for scene in state.scene_list:
        idx = scene.scene_number - 1
        ent = state.entities.get(idx, {})
        # Also try string key fallback
        if not ent:
            ent = state.entities.get(str(idx), {})
        writer.writerow([
            scene.scene_number, scene.heading, scene.int_ext,
            scene.location, scene.time_of_day,
            scene.start_line + 1, scene.end_line,
            f"{scene.confidence:.0%}",
            "是" if scene.is_manually_adjusted else "否",
            scene.summary,
            ", ".join(ent.get("characters", [])),
            ", ".join(ent.get("props", [])),
        ])

    content = "\ufeff" + output.getvalue()  # UTF-8 BOM for Excel
    filename = os.path.splitext(state.filename or "export")[0] + "_scenes.csv"
    return Response(
        content.encode("utf-8"),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": _make_content_disposition(filename)},
    )


@app.route("/api/export/xlsx")
def export_xlsx():
    """Export scenes to a multi-sheet Excel workbook (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "缺少 openpyxl，请运行: pip install openpyxl"}), 500

    state = AppState()
    if not state.scene_list:
        return jsonify({"error": "无场次数据可导出"}), 400

    include_chars    = request.args.get("chars",    "0") == "1"
    include_props    = request.args.get("props",    "0") == "1"
    include_entities = request.args.get("entities", "0") == "1"
    include_colors   = request.args.get("colors",   "0") == "1"

    wb = openpyxl.Workbook()

    # ── Shared styles ─────────────────────────────────────────────────
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    HDR_FILL  = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALT_FILL  = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    WRAP      = Alignment(wrap_text=True, vertical="top")

    def apply_header(ws, headers, col_widths=None):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = HDR_ALIGN
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22
        if col_widths:
            for j, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(j)].width = w

    def _ent(idx):
        """Return entity dict for scene index (int or str key)."""
        e = state.entities.get(idx)
        return e if e is not None else (state.entities.get(str(idx)) or {})

    # ── Sheet 1: 场次列表 (always) ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "场次列表"
    apply_header(ws1,
        ["场次号", "标题", "内外景", "地点", "时间",
         "起始行", "结束行", "置信度", "人工校准", "摘要", "人物", "道具"],
        [7, 32, 7, 20, 7, 7, 7, 7, 7, 40, 24, 24])

    for i, scene in enumerate(state.scene_list):
        e = _ent(i)
        ws1.append([
            scene.scene_number,
            scene.heading,
            scene.int_ext or "",
            scene.location or "",
            scene.time_of_day or "",
            scene.start_line + 1,
            scene.end_line,
            f"{scene.confidence:.0%}",
            "是" if scene.is_manually_adjusted else "否",
            scene.summary or "",
            "、".join(e.get("characters", [])),
            "、".join(e.get("props", [])),
        ])
        r = ws1.max_row
        if i % 2 == 1:
            for cell in ws1[r]:
                cell.fill = ALT_FILL
        # Wrap-text for summary, characters, props columns (J, K, L)
        for cell in list(ws1[r])[9:]:
            cell.alignment = WRAP

    # ── Sheet 2: 人物汇总 ────────────────────────────────────────────
    if include_chars:
        ws2 = wb.create_sheet("人物汇总")
        apply_header(ws2, ["人物名称", "出现场次数", "出现场次列表"], [20, 10, 55])
        char_map: dict = {}
        for i, scene in enumerate(state.scene_list):
            for name in _ent(i).get("characters", []):
                char_map.setdefault(name, []).append(scene.scene_number)
        for ri, (name, sns) in enumerate(sorted(char_map.items())):
            ws2.append([name, len(sns), "、".join(str(s) for s in sns)])
            if ri % 2 == 1:
                for cell in ws2[ws2.max_row]:
                    cell.fill = ALT_FILL

    # ── Sheet 3: 道具汇总 ────────────────────────────────────────────
    if include_props:
        ws3 = wb.create_sheet("道具汇总")
        apply_header(ws3, ["道具名称", "出现场次数", "出现场次列表"], [20, 10, 55])
        prop_map: dict = {}
        for i, scene in enumerate(state.scene_list):
            for name in _ent(i).get("props", []):
                prop_map.setdefault(name, []).append(scene.scene_number)
        for ri, (name, sns) in enumerate(sorted(prop_map.items())):
            ws3.append([name, len(sns), "、".join(str(s) for s in sns)])
            if ri % 2 == 1:
                for cell in ws3[ws3.max_row]:
                    cell.fill = ALT_FILL

    # ── Sheet 4: 场景实体明细 ─────────────────────────────────────────
    if include_entities:
        ws4 = wb.create_sheet("场景实体明细")
        apply_header(ws4,
            ["场次号", "标题", "人物", "道具", "场景类型"],
            [7, 30, 30, 30, 14])
        for i, scene in enumerate(state.scene_list):
            e = _ent(i)
            ws4.append([
                scene.scene_number,
                scene.heading,
                "、".join(e.get("characters", [])),
                "、".join(e.get("props", [])),
                e.get("scene_type", ""),
            ])
            if i % 2 == 1:
                for cell in ws4[ws4.max_row]:
                    cell.fill = ALT_FILL

    # ── Sheet 5: 场景色谱 ─────────────────────────────────────────────
    if include_colors:
        ws5 = wb.create_sheet("场景色谱")
        apply_header(ws5,
            ["地点", "内外景", "时间", "场次号", "标题", "人物", "道具"],
            [20, 8, 8, 7, 30, 24, 24])
        # Sort by location, int_ext, time_of_day for shooting schedule view
        sorted_idx = sorted(range(len(state.scene_list)), key=lambda i: (
            state.scene_list[i].location or "",
            state.scene_list[i].int_ext or "",
            state.scene_list[i].time_of_day or "",
            state.scene_list[i].scene_number,
        ))
        palette = ["EFF6FF", "F0FDF4", "FFF7ED", "FDF4FF", "FFFBEB",
                   "F0F9FF", "FFF1F2", "F5F3FF", "FEFCE8", "F0FDFA"]
        loc_color: dict = {}
        color_idx = 0
        for i in sorted_idx:
            scene = state.scene_list[i]
            e = _ent(i)
            loc = scene.location or ""
            if loc not in loc_color:
                loc_color[loc] = palette[color_idx % len(palette)]
                color_idx += 1
            ws5.append([
                scene.location or "",
                scene.int_ext or "",
                scene.time_of_day or "",
                scene.scene_number,
                scene.heading,
                "、".join(e.get("characters", [])),
                "、".join(e.get("props", [])),
            ])
            fc = loc_color[loc]
            row_fill = PatternFill(start_color=fc, end_color=fc, fill_type="solid")
            for cell in ws5[ws5.max_row]:
                cell.fill = row_fill

    # ── Serialize ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = os.path.splitext(state.filename or "export")[0] + "_breakdown.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _make_content_disposition(filename)},
    )


def _make_content_disposition(filename: str) -> str:
    """Build a Content-Disposition header that supports Unicode filenames (RFC 5987)."""
    encoded = urllib.parse.quote(filename, safe="")
    # Provide ASCII fallback ("export") and RFC 5987 UTF-8 encoded name
    return f"attachment; filename=\"export\"; filename*=UTF-8''{encoded}"


# ── Schedule (排期) ───────────────────────────────────────────────

def _schedule_state_dict(state: AppState) -> dict:
    """Serialize schedule state for the frontend."""
    schedule_dict = state.schedule.to_dict() if state.schedule else None
    violations = []
    if state.schedule and state.scene_list:
        entities_by_num = _entities_by_scene_number(state)
        checker = ConstraintChecker(
            list(state.scene_list),
            entities_by_num,
            state.schedule.actors,
            state.schedule.locations,
        )
        raw = checker.check_all(state.schedule)
        violations = [
            {"type": v.type, "severity": v.severity,
             "message": v.message, "day_date": v.day_date,
             "scene_ids": v.scene_ids}
            for v in raw
        ]
    return {
        "schedule": schedule_dict,
        "snapshots": [s.to_dict() for s in state.schedule_snapshots],
        "violations": violations,
        "task": {
            "status": state.schedule_task.status,
            "error": state.schedule_task.error,
        },
    }


def _entities_by_scene_number(state: AppState) -> dict:
    """Convert AppState.entities (indexed by scene list index) to scene_number keyed dict."""
    result: dict = {}
    if not state.scene_list:
        return result
    scenes = list(state.scene_list)
    for idx, ent in state.entities.items():
        idx_int = int(idx) if isinstance(idx, str) else idx
        if 0 <= idx_int < len(scenes):
            result[scenes[idx_int].scene_number] = ent
    return result


def _save_schedule_snapshot(state: AppState, trigger: str) -> None:
    """Save a snapshot of the current schedule (in-memory + project disk)."""
    if not state.schedule:
        return
    version = (state.schedule_snapshots[-1].version + 1) if state.schedule_snapshots else 1
    snap = ScheduleSnapshot(
        version=version,
        timestamp=datetime.now().isoformat(),
        trigger=trigger,
        schedule_data=state.schedule.to_dict(),
    )
    state.schedule_snapshots.append(snap)
    # Keep at most 20 snapshots in memory
    if len(state.schedule_snapshots) > 20:
        state.schedule_snapshots = state.schedule_snapshots[-20:]
    # Persist to project snapshots/ directory
    if state.project:
        try:
            state.project.save_schedule_snapshot(snap)
        except Exception:
            pass


def _persist_schedule(state: AppState) -> None:
    """Save schedule to project dir if a project is open."""
    if state.project and state.schedule:
        state.project.save_schedule(state.schedule)


def _get_learner(state: AppState) -> "SchedulePreferenceLearner | None":
    """Return the lazily-initialised SchedulePreferenceLearner, or None if no project."""
    if not state.project:
        return None
    if state.preference_learner is None:
        save_path = os.path.join(state.project.project_dir, "schedule_preferences.json")
        state.preference_learner = SchedulePreferenceLearner(save_path)
    return state.preference_learner


def _record_schedule_adjustment(
    state: AppState,
    before: "ProductionSchedule",
    after: "ProductionSchedule",
) -> None:
    """Record a manual schedule adjustment for preference learning (best-effort)."""
    try:
        learner = _get_learner(state)
        if learner is None:
            return
        all_scenes, entities_by_num = _collect_all_scenes(state)
        evaluator = ScheduleEvaluator(
            scenes=all_scenes,
            entities=entities_by_num,
            config=after.config,
        )
        learner.record_adjustment(before, after, evaluator)
    except Exception:
        pass  # never let learning errors surface to the user


@app.route("/schedule")
def schedule_page():
    return render_template("schedule.html")


@app.route("/api/schedule/state")
def schedule_state():
    state = AppState()
    # Auto-load from project on first access
    if state.schedule is None and state.project:
        loaded = state.project.load_schedule()
        if loaded is not None:
            state.schedule = loaded
        # Load snapshots (metadata only for listing; full data loaded on rollback)
        if not state.schedule_snapshots:
            snaps_meta = state.project.list_schedule_snapshots()
            state.schedule_snapshots = []
            for meta in snaps_meta:
                try:
                    snap = state.project.load_schedule_snapshot(meta["version"])
                    state.schedule_snapshots.append(snap)
                except Exception:
                    pass
    return jsonify(_schedule_state_dict(state))


@app.route("/api/schedule/config", methods=["GET", "PUT"])
def schedule_config():
    state = AppState()
    if request.method == "GET":
        cfg = state.schedule.config.to_dict() if state.schedule else ScheduleConfig(
            start_date="").to_dict()
        return jsonify(cfg)

    data = request.get_json() or {}
    if state.schedule is None:
        state.schedule = ProductionSchedule(
            config=ScheduleConfig.from_dict(data),
            start_date=data.get("start_date", ""),
        )
    else:
        state.schedule.config = ScheduleConfig.from_dict(data)
    _persist_schedule(state)
    return jsonify({"ok": True, "config": state.schedule.config.to_dict()})


@app.route("/api/schedule/actors", methods=["POST"])
def schedule_actors():
    state = AppState()
    if state.schedule is None:
        state.schedule = ProductionSchedule(config=ScheduleConfig(start_date=""))

    data = request.get_json() or {}
    actor = ActorSchedule.from_dict(data)
    # Replace if exists, otherwise append
    state.schedule.actors = [
        a for a in state.schedule.actors if a.character_name != actor.character_name
    ]
    state.schedule.actors.append(actor)
    _persist_schedule(state)
    return jsonify(actor.to_dict())


@app.route("/api/schedule/actors/<path:char_name>", methods=["DELETE"])
def schedule_actor(char_name):
    state = AppState()
    if state.schedule is None:
        return jsonify({"error": "暂无排期数据"}), 404

    state.schedule.actors = [
        a for a in state.schedule.actors if a.character_name != char_name
    ]
    _persist_schedule(state)
    return jsonify({"ok": True})


@app.route("/api/schedule/locations", methods=["POST"])
def schedule_locations():
    state = AppState()
    if state.schedule is None:
        state.schedule = ProductionSchedule(config=ScheduleConfig(start_date=""))

    data = request.get_json() or {}
    loc = LocationInfo.from_dict(data)
    state.schedule.locations = [
        l for l in state.schedule.locations if l.name != loc.name
    ]
    state.schedule.locations.append(loc)
    _persist_schedule(state)
    return jsonify(loc.to_dict())


@app.route("/api/schedule/locations/<path:loc_name>", methods=["DELETE"])
def schedule_location(loc_name):
    state = AppState()
    if state.schedule is None:
        return jsonify({"error": "暂无排期数据"}), 404

    state.schedule.locations = [
        l for l in state.schedule.locations if l.name != loc_name
    ]
    _persist_schedule(state)
    return jsonify({"ok": True})



@app.route("/api/schedule/day/<date_str>", methods=["PUT"])
def schedule_update_day(date_str):
    state = AppState()
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 404

    day = next((d for d in state.schedule.shooting_days if d.date == date_str), None)
    if day is None:
        return jsonify({"error": f"找不到拍摄日: {date_str}"}), 404

    import copy as _copy
    before_schedule = _copy.deepcopy(state.schedule)
    _save_schedule_snapshot(state, "manual_adjust")
    data = request.get_json() or {}
    if "scene_ids" in data:
        day.scene_ids = data["scene_ids"]
    if "location" in data:
        day.location = data["location"]
    if "call_time" in data:
        day.call_time = data["call_time"]
    if "estimated_end" in data:
        day.estimated_end = data["estimated_end"]
    if "notes" in data:
        day.notes = data["notes"]
    if "status" in data:
        day.status = data["status"]
    if "weather_backup" in data:
        day.weather_backup = data["weather_backup"]

    _persist_schedule(state)
    _record_schedule_adjustment(state, before_schedule, state.schedule)
    return jsonify(_schedule_state_dict(state))




@app.route("/api/schedule/rollback", methods=["POST"])
def schedule_rollback():
    state = AppState()
    data = request.get_json() or {}
    version = data.get("version")
    if version is None:
        return jsonify({"error": "缺少 version 参数"}), 400

    snap = next((s for s in state.schedule_snapshots if s.version == version), None)
    # Fall back to loading from project disk if not in memory
    if snap is None and state.project:
        try:
            snap = state.project.load_schedule_snapshot(version)
        except FileNotFoundError:
            pass
    if snap is None:
        return jsonify({"error": f"快照版本 {version} 不存在"}), 404

    _save_schedule_snapshot(state, "before_rollback")
    state.schedule = ProductionSchedule.from_dict(snap.schedule_data)
    _persist_schedule(state)
    return jsonify(_schedule_state_dict(state))


@app.route("/api/schedule/llm-advise", methods=["POST"])
def schedule_llm_advise():
    """LLM advisor layer: send schedule summary to LLM for suggestions."""
    state = AppState()
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 400
    if not state.llm_config.provider:
        return jsonify({"error": "请先配置 LLM 设置"}), 400
    if state.schedule_task.status == "running":
        return jsonify({"error": "已有任务运行中"}), 409

    state.schedule_task.status = "running"
    state.schedule_task.error = ""
    state.schedule_task.result = None

    schedule_summary = json.dumps(state.schedule.to_dict(), ensure_ascii=False, indent=2)

    def _run():
        try:
            llm = _create_llm(state.llm_config)
            prompt = (
                "你是一位经验丰富的制片人。以下是当前拍摄排期（JSON），请以制片人视角"
                "给出 3~5 条具体可行的优化建议，重点关注：转场次数、演员成本、均衡性、"
                "是否存在风险。请用中文列表回答，每条建议单独一行。\n\n"
                f"```json\n{schedule_summary}\n```"
            )
            suggestion = llm.complete(prompt)
            state.schedule_task.status = "done"
            state.schedule_task.result = suggestion
        except Exception as e:
            state.schedule_task.status = "error"
            state.schedule_task.error = str(e)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "status": "running"})


@app.route("/api/schedule/llm-status")
def schedule_llm_status():
    state = AppState()
    return jsonify({
        "status": state.schedule_task.status,
        "result": state.schedule_task.result,
        "error": state.schedule_task.error,
    })


# ── Schedule LLM Advisor ─────────────────────────────────────────


def _build_schedule_summary(state: AppState) -> str:
    """Build a human-readable schedule summary for the LLM prompt."""
    if not state.schedule:
        return "（无排期数据）"

    sched = state.schedule
    all_scenes, entities_by_num = _collect_all_scenes(state)
    scene_map = {s.scene_number: s for s in all_scenes}

    lines = []
    lines.append(
        f"开机: {sched.start_date or '未设置'}  "
        f"杀青: {sched.end_date or '未知'}  "
        f"拍摄天数: {len(sched.shooting_days)}  "
        f"总场次: {sum(len(d.scene_ids) for d in sched.shooting_days)}"
    )
    lines.append("")

    weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    for day in sched.shooting_days:
        try:
            import datetime as _dt
            dt = _dt.date.fromisoformat(day.date)
            wd = weekdays[dt.weekday()]
        except Exception:
            wd = ""

        actors: set = set()
        scene_types: list = []
        locs: list = []
        for sid in day.scene_ids:
            sc = scene_map.get(sid)
            if sc:
                if sc.location:
                    locs.append(sc.location)
            ent = entities_by_num.get(sid, {})
            actors.update(ent.get("characters", []))
            st = ent.get("scene_type", "")
            if st:
                scene_types.append(st)

        loc_str = "、".join(dict.fromkeys(locs)) if locs else day.location or "未指定"
        actor_str = "、".join(sorted(actors)) if actors else "无"
        type_str = "、".join(dict.fromkeys(scene_types)) if scene_types else "未分类"

        lines.append(
            f"第{day.day_number}天 ({day.date} {wd})  "
            f"地点: {loc_str}  "
            f"场次: {day.scene_ids}  "
            f"演员: {actor_str}  "
            f"场景类型: {type_str}"
        )

    return "\n".join(lines)


def _build_constraints_summary(state: AppState, conflicts: list) -> str:
    """Build a human-readable constraints + conflicts summary for the LLM prompt."""
    lines = []

    if state.schedule:
        for actor in state.schedule.actors:
            parts = []
            if actor.actor_name:
                parts.append(f"演员: {actor.actor_name}")
            if actor.unavailable_dates:
                parts.append(f"不可用日期: {actor.unavailable_dates}")
            if actor.daily_rate:
                parts.append(f"日薪: {actor.daily_rate}元")
            if parts:
                lines.append(f"· {actor.character_name} — " + "，".join(parts))

        for loc in state.schedule.locations:
            parts = []
            if loc.cost_per_day:
                parts.append(f"费用 {loc.cost_per_day}元/天")
            if loc.travel_time_minutes:
                parts.append(f"转场时间 {loc.travel_time_minutes}分钟")
            if loc.available_dates:
                parts.append(f"可用日期: {loc.available_dates[:5]}")
            if parts:
                lines.append(f"· 场地「{loc.name}」— " + "，".join(parts))

    if conflicts:
        lines.append("")
        lines.append(f"当前冲突 ({len(conflicts)} 条):")
        for c in conflicts[:10]:
            lines.append(f"  ⚠ [{c.get('severity','error')}] {c.get('message','')}"
                         + (f" (日期: {c.get('day_date','')})" if c.get("day_date") else ""))
    else:
        lines.append("当前无硬约束冲突。")

    return "\n".join(lines) if lines else "无约束信息。"


def _build_day_scenes_text(state: AppState, day) -> str:
    """Build a readable description of all scenes in a shooting day."""
    all_scenes, entities_by_num = _collect_all_scenes(state)
    scene_map = {s.scene_number: s for s in all_scenes}
    lines = []

    for sid in day.scene_ids:
        sc = scene_map.get(sid)
        ent = entities_by_num.get(sid, {})
        actors = "、".join(ent.get("characters", [])) or "无"
        props  = "、".join(ent.get("props", [])) or "无"
        stype  = ent.get("scene_type", "") or "未分类"

        if sc:
            loc_info = " · ".join(filter(None, [sc.int_ext, sc.location, sc.time_of_day]))
            lines.append(
                f"【场次 {sid}】{sc.heading or '(无标题)'}\n"
                f"  位置: {loc_info or '未知'}\n"
                f"  类型: {stype}  演员: {actors}  道具: {props}\n"
                + (f"  摘要: {sc.summary}\n" if sc.summary else "")
            )
        else:
            lines.append(f"【场次 {sid}】（无详细信息）  演员: {actors}")

    return "\n".join(lines) if lines else "（该拍摄日无场次数据）"


@app.route("/api/schedule/llm/analyze", methods=["POST"])
def schedule_llm_analyze():
    """Async: LLM analyzes the full schedule from a Line Producer perspective.

    Uses state.llm_task for tracking — poll GET /api/llm/status.
    """
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 400
    if not state.llm_config.api_key:
        return jsonify({"error": "请先配置 LLM API Key"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有任务运行中，请稍后"}), 409

    # Build summaries synchronously before handing off to thread
    from src.schedule.constraints import ConstraintChecker
    all_scenes, entities_by_num = _collect_all_scenes(state)
    checker = ConstraintChecker(
        all_scenes, entities_by_num,
        state.schedule.actors, state.schedule.locations,
    )
    conflicts_raw = checker.check_all(state.schedule)
    conflicts = [{"severity": v.severity, "message": v.message, "day_date": v.day_date}
                 for v in conflicts_raw]

    schedule_summary    = _build_schedule_summary(state)
    constraints_summary = _build_constraints_summary(state, conflicts)

    state.llm_task.task_type = "schedule_analyze"
    state.llm_task.status    = "running"
    state.llm_task.error     = ""
    state.llm_task.result    = None

    def _run():
        try:
            from src.schedule.llm_advisor import LLMScheduleAdvisor
            advisor = LLMScheduleAdvisor(_create_llm(state.llm_config))
            result  = advisor.analyze_schedule(schedule_summary, constraints_summary)
            state.llm_task.status = "done"
            state.llm_task.result = result
        except Exception as exc:
            state.llm_task.status = "error"
            state.llm_task.error  = str(exc)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "status": "running"})


@app.route("/api/schedule/llm/suggest_day/<date>", methods=["POST"])
def schedule_llm_suggest_day(date):
    """Async: LLM suggests optimal shooting order for scenes on a given day.

    Uses state.llm_task for tracking — poll GET /api/llm/status.
    """
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 400
    if not state.llm_config.api_key:
        return jsonify({"error": "请先配置 LLM API Key"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有任务运行中，请稍后"}), 409

    day = next((d for d in state.schedule.shooting_days if d.date == date), None)
    if day is None:
        return jsonify({"error": f"找不到拍摄日: {date}"}), 404
    if not day.scene_ids:
        return jsonify({"error": "该拍摄日暂无场次"}), 400

    scenes_text = _build_day_scenes_text(state, day)

    state.llm_task.task_type = "schedule_suggest_day"
    state.llm_task.status    = "running"
    state.llm_task.error     = ""
    state.llm_task.result    = None

    def _run():
        try:
            from src.schedule.llm_advisor import LLMScheduleAdvisor
            advisor = LLMScheduleAdvisor(_create_llm(state.llm_config))
            result  = advisor.suggest_day_arrangement(scenes_text)
            state.llm_task.status = "done"
            state.llm_task.result = result
        except Exception as exc:
            state.llm_task.status = "error"
            state.llm_task.error  = str(exc)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "status": "running"})


@app.route("/api/schedule/llm/compare", methods=["POST"])
def schedule_llm_compare():
    """Async: LLM compares two schedule options and recommends one.

    Body:
      option_a : str  — human-readable description of option A
                        (if omitted, the current schedule summary is used)
      option_b : str  — human-readable description of option B (required)

    Poll GET /api/llm/status for completion; result is the comparison text.
    """
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 400
    if not state.llm_config.api_key:
        return jsonify({"error": "请先配置 LLM API Key"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有任务运行中，请稍后"}), 409

    data = request.get_json() or {}
    option_b = data.get("option_b", "").strip()
    if not option_b:
        return jsonify({"error": "缺少 option_b 字段"}), 400

    # option_a defaults to the current schedule if not provided
    option_a = data.get("option_a", "").strip() or _build_schedule_summary(state)

    state.llm_task.task_type = "schedule_compare"
    state.llm_task.status    = "running"
    state.llm_task.error     = ""
    state.llm_task.result    = None

    def _run():
        try:
            from src.schedule.llm_advisor import LLMScheduleAdvisor
            advisor = LLMScheduleAdvisor(_create_llm(state.llm_config))
            result  = advisor.evaluate_trade_off(option_a, option_b)
            state.llm_task.status = "done"
            state.llm_task.result = result
        except Exception as exc:
            state.llm_task.status = "error"
            state.llm_task.error  = str(exc)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "status": "running"})


@app.route("/api/schedule/llm/contingency", methods=["POST"])
def schedule_llm_contingency():
    """Async: LLM generates a contingency plan for a specified risk scenario.

    Body:
      risk_type : str  — one of: "rain" | "actor_sick" |
                         "location_unavailable" | "overtime"
                         (any other string is passed through as a custom risk label)

    Poll GET /api/llm/status for completion; result is the contingency plan text.
    """
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 400
    if not state.llm_config.api_key:
        return jsonify({"error": "请先配置 LLM API Key"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有任务运行中，请稍后"}), 409

    data = request.get_json() or {}
    risk_type = data.get("risk_type", "").strip()
    if not risk_type:
        return jsonify({"error": "缺少 risk_type 字段"}), 400

    schedule_summary = _build_schedule_summary(state)

    state.llm_task.task_type = "schedule_contingency"
    state.llm_task.status    = "running"
    state.llm_task.error     = ""
    state.llm_task.result    = None

    def _run():
        try:
            from src.schedule.llm_advisor import LLMScheduleAdvisor
            advisor = LLMScheduleAdvisor(_create_llm(state.llm_config))
            result  = advisor.suggest_contingency(schedule_summary, risk_type)
            state.llm_task.status = "done"
            state.llm_task.result = result
        except Exception as exc:
            state.llm_task.status = "error"
            state.llm_task.error  = str(exc)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "status": "running", "risk_type": risk_type})


@app.route("/api/schedule/llm/actor-workload", methods=["POST"])
def schedule_llm_actor_workload():
    """Async: LLM analyses a specific actor's workload and flags over-scheduling risks.

    Body:
      actor_name : str  — character name (角色名) to analyse

    Poll GET /api/llm/status for completion; result is the workload analysis text.
    """
    state = AppState()
    _ensure_schedule_loaded(state)
    if not state.schedule:
        return jsonify({"error": "暂无排期数据"}), 400
    if not state.llm_config.api_key:
        return jsonify({"error": "请先配置 LLM API Key"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有任务运行中，请稍后"}), 409

    data = request.get_json() or {}
    actor_name = data.get("actor_name", "").strip()
    if not actor_name:
        return jsonify({"error": "缺少 actor_name 字段"}), 400

    schedule_summary = _build_schedule_summary(state)

    state.llm_task.task_type = "schedule_actor_workload"
    state.llm_task.status    = "running"
    state.llm_task.error     = ""
    state.llm_task.result    = None

    def _run():
        try:
            from src.schedule.llm_advisor import LLMScheduleAdvisor
            advisor = LLMScheduleAdvisor(_create_llm(state.llm_config))
            result  = advisor.analyze_actor_workload(schedule_summary, actor_name)
            state.llm_task.status = "done"
            state.llm_task.result = result
        except Exception as exc:
            state.llm_task.status = "error"
            state.llm_task.error  = str(exc)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "status": "running", "actor_name": actor_name})


# ── Callsheet Page ────────────────────────────────────────────────


@app.route("/callsheet")
def callsheet_page():
    """Serve the callsheet management page."""
    return render_template("callsheet.html")


# ── Callsheet (通告单) ────────────────────────────────────────────


@app.route("/api/callsheets", methods=["GET"])
def list_callsheets():
    """Return summary list of all saved callsheets for the current project."""
    state = AppState()
    if not state.project:
        return jsonify([])
    return jsonify(state.project.list_callsheets())


@app.route("/api/callsheets/<date>", methods=["GET"])
def get_callsheet(date: str):
    """Return a saved callsheet by date (YYYY-MM-DD)."""
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400
    try:
        cs = state.project.load_callsheet(date)
        return jsonify(cs.to_dict())
    except FileNotFoundError:
        return jsonify({"error": f"通告单不存在: {date}"}), 404


@app.route("/api/callsheets/<date>/generate", methods=["POST"])
def generate_callsheet(date: str):
    """Generate (and save) a callsheet for the given shooting day."""
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    _ensure_schedule_loaded(state)
    if state.schedule is None:
        return jsonify({"error": "尚未生成排期"}), 400

    day = next((d for d in state.schedule.shooting_days if d.date == date), None)
    if day is None:
        return jsonify({"error": f"找不到拍摄日: {date}"}), 404

    all_scenes, entities_by_num = _collect_all_scenes(state)

    from src.callsheet.generator import CallSheetGenerator
    gen = CallSheetGenerator()
    cs = gen.generate(
        shooting_day=day,
        scenes=all_scenes,
        entities=entities_by_num,
        actors=state.schedule.actors,
        schedule=state.schedule,
        project_name=state.project.meta.name,
    )
    state.project.save_callsheet(cs)
    return jsonify(cs.to_dict())


@app.route("/api/callsheets/<date>", methods=["PUT"])
def save_callsheet(date: str):
    """Save (overwrite) a callsheet with the provided JSON body."""
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    data = request.get_json(silent=True) or {}
    data["date"] = date   # ensure date is consistent

    from src.callsheet.models import CallSheet
    cs = CallSheet.from_dict(data)
    state.project.save_callsheet(cs)
    return jsonify({"ok": True})


@app.route("/api/callsheets/<date>", methods=["DELETE"])
def delete_callsheet(date: str):
    """Delete a saved callsheet file."""
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    import os as _os
    path = _os.path.join(state.project.project_dir, "callsheets", f"{date}.json")
    if not _os.path.exists(path):
        return jsonify({"error": f"通告单不存在: {date}"}), 404
    _os.remove(path)
    return jsonify({"ok": True})


@app.route("/api/callsheets/<date>/export/xlsx", methods=["GET"])
def export_callsheet_xlsx(date: str):
    """Export a saved callsheet as a styled .xlsx file."""
    state = AppState()
    if not state.project:
        return jsonify({"error": "未打开项目"}), 400

    try:
        import openpyxl  # noqa: F401 — verify availability
    except ImportError:
        return jsonify({"error": "缺少 openpyxl，请运行: pip install openpyxl"}), 500

    try:
        cs = state.project.load_callsheet(date)
    except FileNotFoundError:
        return jsonify({"error": f"通告单不存在: {date}"}), 404

    from src.callsheet.exporter import CallSheetExporter
    xlsx_bytes = CallSheetExporter().export_xlsx(cs)

    filename = f"通告单_{date}.xlsx"
    return Response(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"},
    )


@app.route("/api/callsheets/<date>/export/pdf", methods=["GET"])
def export_callsheet_pdf(date: str):
    """Return a print-ready HTML page for the callsheet (browser print → PDF)."""
    state = AppState()
    if not state.project:
        return "<h3>未打开项目</h3>", 400

    try:
        cs = state.project.load_callsheet(date)
    except FileNotFoundError:
        return f"<h3>通告单不存在: {date}</h3>", 404

    from src.callsheet.exporter import CallSheetExporter
    html = CallSheetExporter().export_pdf_html(cs)
    return Response(html, mimetype="text/html; charset=utf-8")


@app.route("/api/callsheets/<date>/llm/notes", methods=["POST"])
def callsheet_llm_notes(date: str):
    """Async: analyse scene content for the shooting day and generate production notes."""
    state = AppState()
    if not state.llm_config.api_key:
        return jsonify({"error": "请先配置 LLM API Key"}), 400
    if state.llm_task.status == "running":
        return jsonify({"error": "已有任务运行中，请稍后"}), 409

    _ensure_schedule_loaded(state)
    if state.schedule is None:
        return jsonify({"error": "尚未生成排期"}), 400

    day = next((d for d in state.schedule.shooting_days if d.date == date), None)
    if day is None:
        return jsonify({"error": f"找不到拍摄日: {date}"}), 404
    if not day.scene_ids:
        return jsonify({"error": "该拍摄日暂无场次"}), 400

    # Collect raw scene content for scenes in this day
    target_ids = set(day.scene_ids)
    parts: list[str] = []

    if state.project:
        state.save_to_episode()
        global_num = 1
        for ep_id in state.project.meta.episode_order:
            try:
                ep = state.project.load_episode(ep_id)
            except Exception:
                continue
            for sd in (ep.scenes or []):
                if global_num in target_ids:
                    heading = sd.get("heading", "")
                    content = sd.get("content", "") or sd.get("summary", "")
                    parts.append(f"【场次{global_num}】{heading}\n{content}")
                global_num += 1
    elif state.scene_list:
        for scene in state.scene_list:
            if scene.scene_number in target_ids:
                text_body = getattr(scene, "content", "") or getattr(scene, "summary", "")
                parts.append(f"【场次{scene.scene_number}】{scene.heading}\n{text_body}")

    if not parts:
        return jsonify({"error": "未找到场次内容，请确认已完成剧本拆解"}), 400

    scene_content = "\n\n---\n\n".join(parts)

    state.llm_task.task_type = "callsheet_notes"
    state.llm_task.status    = "running"
    state.llm_task.error     = ""
    state.llm_task.result    = None

    def _run():
        try:
            from src.schedule.llm_advisor import LLMScheduleAdvisor
            advisor = LLMScheduleAdvisor(_create_llm(state.llm_config))
            result  = advisor.generate_notes(scene_content)
            state.llm_task.status = "done"
            state.llm_task.result = result
        except Exception as exc:
            state.llm_task.status = "error"
            state.llm_task.error  = str(exc)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "status": "running"})


# ── Schedule Preference Learner API ──────────────────────────────────────────

@app.route("/api/schedule/learned-preferences", methods=["GET"])
def get_learned_preferences():
    """Return the current learned weights and rules."""
    state = AppState()
    learner = _get_learner(state)
    if learner is None:
        return jsonify({"error": "未打开项目，偏好数据不可用"}), 400
    return jsonify(learner.to_dict())


@app.route("/api/schedule/learned-preferences/reset", methods=["POST"])
def reset_learned_preferences():
    """Reset all learned weights and rules to factory defaults."""
    state = AppState()
    learner = _get_learner(state)
    if learner is None:
        return jsonify({"error": "未打开项目，偏好数据不可用"}), 400
    learner.reset()
    return jsonify({"ok": True, "message": "已重置学习数据"})


@app.route("/api/schedule/learned-preferences/add-rule", methods=["POST"])
def add_learned_rule():
    """Manually add a structural rule to the preference learner.

    Body: { "type": str, "scenes": [int, ...], "params": {}, "reason": str }
    """
    state = AppState()
    learner = _get_learner(state)
    if learner is None:
        return jsonify({"error": "未打开项目，偏好数据不可用"}), 400
    rule = request.get_json() or {}
    if not rule.get("type") or not rule.get("scenes"):
        return jsonify({"error": "缺少 type 或 scenes 字段"}), 400
    learner.record_rule_from_llm(rule)
    return jsonify({"ok": True, "rules": learner.learned_rules})


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5001, debug=True)
