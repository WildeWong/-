"""CallSheetExporter — exports a CallSheet to XLSX or print-ready HTML."""
from __future__ import annotations

import io
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from .models import CallSheet


class CallSheetExporter:
    """Export a CallSheet to different file formats."""

    # ── XLSX ──────────────────────────────────────────────────────

    def export_xlsx(self, callsheet: "CallSheet") -> bytes:
        """Generate a styled XLSX callsheet using openpyxl.

        Returns:
            Raw bytes of the .xlsx file.
        Raises:
            ImportError: if openpyxl is not installed.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "通告单"

        # ── Shared styles ──────────────────────────────────────────
        BLUE       = "2563EB"
        BLUE_LIGHT = "EFF6FF"
        GRAY_HDR   = "F1F5F9"
        SECTION_BG = "1E293B"
        AMBER      = "FEF3C7"

        def _font(bold=False, size=11, color="000000", italic=False):
            return Font(bold=bold, size=size, color=color, italic=italic,
                        name="PingFang SC")

        def _fill(hex_color):
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        def _align(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def _border_thin():
            thin = Side(style="thin", color="CBD5E1")
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        def _set_row(ws, row, values, font=None, fill=None, align=None, height=None):
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if font:  cell.font = font
                if fill:  cell.fill = fill
                if align: cell.alignment = align
                cell.border = _border_thin()
            if height:
                ws.row_dimensions[row].height = height

        # ── Title row ─────────────────────────────────────────────
        r = 1
        ws.merge_cells(f"A{r}:H{r}")
        cell = ws.cell(row=r, column=1,
                       value=f"每日拍摄通告单  ·  {callsheet.date}  第 {callsheet.day_number} 拍摄日")
        cell.font  = _font(bold=True, size=14, color="FFFFFF")
        cell.fill  = _fill(SECTION_BG)
        cell.alignment = _align("center", "center")
        ws.row_dimensions[r].height = 32

        # ── Info block ────────────────────────────────────────────
        r += 1
        info_pairs = [
            ("项目名称", callsheet.project_name or "—"),
            ("日期",     callsheet.date),
            ("导演",     callsheet.director or "—"),
            ("制片",     callsheet.producer or "—"),
            ("全体集合", callsheet.crew_call or "—"),
            ("拍摄地点", callsheet.location or "—"),
        ]
        for i, (label, value) in enumerate(info_pairs):
            col = (i % 3) * 2 + 1
            if i % 3 == 0 and i > 0:
                r += 1
            label_cell = ws.cell(row=r, column=col, value=label)
            label_cell.font = _font(bold=True, size=10, color="64748B")
            label_cell.fill = _fill(GRAY_HDR)
            label_cell.alignment = _align("right", "center")
            label_cell.border = _border_thin()

            val_cell = ws.cell(row=r, column=col + 1, value=value)
            val_cell.font = _font(size=11)
            val_cell.alignment = _align("left", "center")
            val_cell.border = _border_thin()
            ws.row_dimensions[r].height = 20

        # If last info row had < 3 pairs, merge remaining
        r += 1  # blank separator
        ws.row_dimensions[r].height = 6

        # ── Section header helper ─────────────────────────────────
        def section_header(ws, row, title, ncols=8):
            ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
            cell = ws.cell(row=row, column=1, value=title)
            cell.font      = _font(bold=True, size=11, color="FFFFFF")
            cell.fill      = _fill(BLUE)
            cell.alignment = _align("left", "center")
            cell.border    = _border_thin()
            ws.row_dimensions[row].height = 22

        # ── Scene table ───────────────────────────────────────────
        r += 1
        section_header(ws, r, "▌ 场次表", ncols=8)
        r += 1
        scene_headers = ["场次号", "场景标题", "内外景", "时间段", "人物", "页数", "备注", ""]
        _set_row(ws, r, scene_headers,
                 font=_font(bold=True, size=10, color="1E293B"),
                 fill=_fill(BLUE_LIGHT),
                 align=_align("center", "center"),
                 height=20)

        for idx, sc in enumerate(callsheet.scenes):
            r += 1
            row_fill = _fill("FFFFFF") if idx % 2 == 0 else _fill("F8FAFC")
            _set_row(ws, r, [
                sc.scene_number,
                sc.heading,
                sc.int_ext,
                sc.time_of_day,
                "、".join(sc.cast_ids),
                f"{sc.pages:.1f}P" if sc.pages else "",
                sc.notes,
                "",
            ], font=_font(size=11), fill=row_fill,
               align=_align("left", "center", wrap=True), height=18)
            ws.cell(row=r, column=1).alignment = _align("center", "center")

        # ── Cast table ────────────────────────────────────────────
        r += 2
        section_header(ws, r, "▌ 演员通告", ncols=8)
        r += 1
        cast_headers = ["角色", "演员", "化妆时间", "到组时间", "上场时间", "状态", "服装备注", ""]
        _set_row(ws, r, cast_headers,
                 font=_font(bold=True, size=10, color="1E293B"),
                 fill=_fill(BLUE_LIGHT),
                 align=_align("center", "center"),
                 height=20)

        for idx, cc in enumerate(callsheet.cast):
            r += 1
            row_fill = _fill("FFFFFF") if idx % 2 == 0 else _fill("F8FAFC")
            status_label = "拍摄" if cc.status == "W" else "备机"
            _set_row(ws, r, [
                cc.character_name,
                cc.actor_name,
                cc.makeup_time,
                cc.call_time,
                cc.on_set_time,
                status_label,
                cc.wardrobe_notes,
                "",
            ], font=_font(size=11), fill=row_fill,
               align=_align("left", "center"), height=18)
            ws.cell(row=r, column=3).alignment = _align("center", "center")
            ws.cell(row=r, column=4).alignment = _align("center", "center")
            ws.cell(row=r, column=5).alignment = _align("center", "center")
            ws.cell(row=r, column=6).alignment = _align("center", "center")

        # ── Props ─────────────────────────────────────────────────
        all_props: set[str] = set()
        for sc in callsheet.scenes:
            all_props.update(sc.props or [])

        if all_props:
            r += 2
            section_header(ws, r, "▌ 道具清单", ncols=8)
            r += 1
            props_text = "、".join(sorted(all_props))
            ws.merge_cells(f"A{r}:H{r}")
            c = ws.cell(row=r, column=1, value=props_text)
            c.font      = _font(size=11)
            c.fill      = _fill(AMBER)
            c.alignment = _align("left", "center", wrap=True)
            c.border    = _border_thin()
            ws.row_dimensions[r].height = max(20, 18 * (1 + len(props_text) // 80))

        # ── Notes ─────────────────────────────────────────────────
        if callsheet.general_notes:
            r += 2
            section_header(ws, r, "▌ 注意事项", ncols=8)
            r += 1
            ws.merge_cells(f"A{r}:H{r}")
            c = ws.cell(row=r, column=1, value=callsheet.general_notes)
            c.font      = _font(size=11)
            c.fill      = _fill("FFFBEB")
            c.alignment = _align("left", "top", wrap=True)
            c.border    = _border_thin()
            lines = callsheet.general_notes.count("\n") + 1
            ws.row_dimensions[r].height = max(20, lines * 15)

        # ── Next day preview ──────────────────────────────────────
        if callsheet.next_day_preview:
            r += 2
            section_header(ws, r, "▌ 明日预告", ncols=8)
            r += 1
            ws.merge_cells(f"A{r}:H{r}")
            c = ws.cell(row=r, column=1, value=callsheet.next_day_preview)
            c.font      = _font(size=11, italic=True)
            c.fill      = _fill(BLUE_LIGHT)
            c.alignment = _align("left", "center", wrap=True)
            c.border    = _border_thin()
            ws.row_dimensions[r].height = 20

        # ── Column widths ─────────────────────────────────────────
        col_widths = [9, 32, 9, 9, 24, 7, 28, 4]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.sheet_view.showGridLines = False

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Print-ready HTML ──────────────────────────────────────────

    def export_pdf_html(self, callsheet: "CallSheet") -> str:
        """Return a print-optimised HTML string.

        The browser can render this page and use File > Print / Ctrl+P to save
        as PDF (no external PDF library required).
        """
        def row(cells, tag="td", extra=""):
            parts = [f"<{tag}{extra}>{c}</{tag}>" for c in cells]
            return "<tr>" + "".join(parts) + "</tr>"

        # Scene rows
        scene_rows = ""
        for sc in callsheet.scenes:
            scene_rows += row([
                sc.scene_number,
                _h(sc.heading),
                sc.int_ext or "",
                sc.time_of_day or "",
                _h("、".join(sc.cast_ids)),
                f"{sc.pages:.1f}P" if sc.pages else "",
                _h(sc.notes),
            ])

        # Cast rows
        cast_rows = ""
        for cc in callsheet.cast:
            cast_rows += row([
                _h(cc.character_name),
                _h(cc.actor_name),
                cc.makeup_time or "—",
                cc.call_time   or "—",
                cc.on_set_time or "—",
                "拍摄" if cc.status == "W" else "备机",
                _h(cc.wardrobe_notes),
            ])

        # Props
        all_props: set[str] = set()
        for sc in callsheet.scenes:
            all_props.update(sc.props or [])
        props_html = ("、".join(sorted(all_props))) if all_props else "无"

        return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>通告单 {_h(callsheet.date)}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: "PingFang SC", "Microsoft YaHei", "Helvetica Neue", Arial, sans-serif;
    font-size: 12px;
    color: #1e293b;
    background: #fff;
    padding: 20px 24px;
  }}
  /* Title */
  .title {{
    font-size: 18px;
    font-weight: 700;
    color: #1e293b;
    text-align: center;
    margin-bottom: 4px;
    letter-spacing: 2px;
  }}
  .subtitle {{
    font-size: 13px;
    color: #64748b;
    text-align: center;
    margin-bottom: 16px;
  }}
  /* Info grid */
  .info-grid {{
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 6px;
    margin-bottom: 16px;
  }}
  .info-cell {{
    display: flex;
    gap: 6px;
    align-items: baseline;
  }}
  .info-label {{
    font-size: 10px;
    color: #64748b;
    font-weight: 600;
    white-space: nowrap;
  }}
  .info-value {{
    font-size: 12px;
    font-weight: 600;
    color: #0f172a;
  }}
  /* Section */
  .section-title {{
    background: #1e293b;
    color: #fff;
    font-size: 11px;
    font-weight: 700;
    padding: 4px 10px;
    letter-spacing: .5px;
    margin: 14px 0 0;
  }}
  /* Tables */
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 11px;
    margin-bottom: 2px;
  }}
  th {{
    background: #eff6ff;
    color: #1e40af;
    font-weight: 700;
    padding: 4px 6px;
    text-align: left;
    border: 1px solid #cbd5e1;
    white-space: nowrap;
  }}
  td {{
    padding: 4px 6px;
    border: 1px solid #e2e8f0;
    vertical-align: middle;
  }}
  tr:nth-child(even) td {{ background: #f8fafc; }}
  /* Props */
  .props-block {{
    background: #fef3c7;
    border: 1px solid #fde68a;
    padding: 6px 10px;
    font-size: 11px;
    line-height: 1.8;
    margin-bottom: 2px;
  }}
  /* Notes */
  .notes-block {{
    background: #fffbeb;
    border: 1px solid #fde68a;
    padding: 8px 10px;
    font-size: 11px;
    white-space: pre-wrap;
    line-height: 1.7;
  }}
  /* Next day */
  .nextday-block {{
    background: #eff6ff;
    border: 1px solid #bfdbfe;
    padding: 6px 10px;
    font-size: 11px;
    color: #1e40af;
    font-style: italic;
  }}
  /* Print */
  @media print {{
    body {{ padding: 10mm 12mm; font-size: 11px; }}
    .no-print {{ display: none; }}
    table {{ page-break-inside: auto; }}
    tr {{ page-break-inside: avoid; }}
  }}
</style>
</head>
<body>

<div class="no-print" style="background:#2563eb;color:#fff;padding:8px 16px;margin:-20px -24px 16px;font-size:13px;display:flex;justify-content:space-between;align-items:center">
  <span>通告单预览 — 使用浏览器打印功能（Ctrl+P）另存为 PDF</span>
  <button onclick="window.print()" style="background:#fff;color:#2563eb;border:none;padding:4px 14px;border-radius:4px;font-weight:600;cursor:pointer">立即打印</button>
</div>

<div class="title">每日拍摄通告单</div>
<div class="subtitle">{_h(callsheet.project_name or "")}　第 {callsheet.day_number} 拍摄日</div>

<div class="info-grid">
  <div class="info-cell"><span class="info-label">日期</span><span class="info-value">{_h(callsheet.date)}</span></div>
  <div class="info-cell"><span class="info-label">全体集合</span><span class="info-value">{_h(callsheet.crew_call)}</span></div>
  <div class="info-cell"><span class="info-label">地点</span><span class="info-value">{_h(callsheet.location)}</span></div>
  <div class="info-cell"><span class="info-label">地址</span><span class="info-value">{_h(callsheet.location_address or "—")}</span></div>
  <div class="info-cell"><span class="info-label">导演</span><span class="info-value">{_h(callsheet.director or "—")}</span></div>
  <div class="info-cell"><span class="info-label">制片</span><span class="info-value">{_h(callsheet.producer or "—")}</span></div>
</div>

<div class="section-title">▌ 场次表</div>
<table>
  <thead>{row(["场次号","场景标题","内外景","时间段","人物","页数","备注"], tag="th")}</thead>
  <tbody>{scene_rows or "<tr><td colspan='7' style='color:#94a3b8;text-align:center'>无场次</td></tr>"}</tbody>
</table>

<div class="section-title">▌ 演员通告</div>
<table>
  <thead>{row(["角色","演员","化妆时间","到组时间","上场时间","状态","服装备注"], tag="th")}</thead>
  <tbody>{cast_rows or "<tr><td colspan='7' style='color:#94a3b8;text-align:center'>无演员</td></tr>"}</tbody>
</table>

<div class="section-title">▌ 道具清单</div>
<div class="props-block">{props_html}</div>

{"<div class='section-title'>▌ 注意事项</div><div class='notes-block'>" + _h(callsheet.general_notes) + "</div>" if callsheet.general_notes else ""}

{"<div class='section-title'>▌ 明日预告</div><div class='nextday-block'>" + _h(callsheet.next_day_preview) + "</div>" if callsheet.next_day_preview else ""}

</body>
</html>"""


def _h(v) -> str:
    """HTML-escape a value."""
    if not v:
        return ""
    return (str(v)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))
